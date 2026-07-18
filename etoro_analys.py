#!/usr/bin/env python3
"""
eToro portföljanalys — konsensus + teknisk analys + analytikerdata
==================================================================

KÖRNING (Terminal på Mac):
  1. Installera beroenden (en gång):
       pip3 install requests yfinance openpyxl pandas anthropic

  2. Lägg dina nycklar i .env-filen i samma mapp (klistra INTE in dem här):
       ETORO_API_KEY=din_offentliga_nyckel
       ETORO_USER_KEY=din_privata_nyckel
       ANTHROPIC_API_KEY=din_anthropic_nyckel   (för Claude-analysen, valfri)

  3. Kör:
       python3 etoro_analys.py

Resultat: portfolj_analys.xlsx i samma mapp (flikar: Konsensus & Analys,
Teknisk analys med Claudes bedömning, Historik med ändringslogg, en flik
per profil). Historiken sparas i portfolj_historik.json mellan körningar.
"""

import os
import sys
import uuid
import json
import requests

# ----------------------------------------------------------------------
# Konfiguration
# ----------------------------------------------------------------------
PROFILES = ["thomaspj", "michalhla", "JeppeKirkBonde", "triangulacapital", "Smudliczek", "ingruc"]

# Konsensuströsklar som ANDEL av signalgruppen (ersätter absoluta MIN_PORTFOLIOS=3)
# med hysteres: högre krav för att komma IN på listan än för att LIGGA KVAR —
# förhindrar flappande IN/UT-poster i historiken när en enda trader trimmar.
# Antal ägare räknas mot ceil(andel × gruppstorlek): 6 profiler → IN 4, KVAR 3.
KONSENSUS_ANDEL_IN = 0.60
KONSENSUS_ANDEL_KVAR = 0.50
KONSENSUS_REGEL = "andel_hysteres_v1"   # versionsmarkör i historiksnapshoten —
# nivåövergångar loggas bara mellan körningar med samma regelversion
OUTPUT_FILE = "portfolj_analys.xlsx"

API_BASE = "https://public-api.etoro.com/api/v1"


def load_env_file():
    """Läs nycklar från .env i skriptets mapp om de inte redan finns i miljön."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


load_env_file()
API_KEY = os.environ.get("ETORO_API_KEY")
USER_KEY = os.environ.get("ETORO_USER_KEY")


def headers():
    return {
        "x-api-key": API_KEY,
        "x-user-key": USER_KEY,
        "x-request-id": str(uuid.uuid4()),
        "Accept": "application/json",
    }


def api_get(path, params=None, quiet=False, retries=3):
    """GET-anrop mot eToro:s publika API med felhantering och 429-respekt."""
    import time
    url = f"{API_BASE}{path}"
    for _ in range(retries):
        r = requests.get(url, headers=headers(), params=params, timeout=30)
        if r.status_code == 429:
            wait = int(r.headers.get("Retry-After", 15))
            print(f"  Rate limit — väntar {wait}s...")
            time.sleep(wait)
            continue
        if not quiet:
            print(f"  GET {path} -> {r.status_code}")
        if r.status_code == 200:
            return r.json()
        print(f"    Svar ({path}): {r.text[:200]}")
        return None
    return None


# ----------------------------------------------------------------------
# Steg 1: Hämta portföljer
# ----------------------------------------------------------------------
_instrument_cache = {}
_industry_cache = {}   # ticker -> stocksIndustryID (1=Råvaror ... 9=Kraft, se app.py)

def resolve_instruments(instrument_ids):
    """Översätt instrument-ID:n till tickers.

    /market-data/instruments utan parametrar returnerar HELA instrumentlistan
    (~15 500 st) med symbolFull — kommaseparerade instrumentIds ger 500, så vi
    hämtar allt en gång och slår upp lokalt.
    """
    if not _instrument_cache:
        print("  Hämtar hela instrumentlistan (en gång)...")
        data = api_get("/market-data/instruments")
        for item in (data or {}).get("instrumentDisplayDatas", []):
            iid = item.get("instrumentID")
            ticker = item.get("symbolFull")
            if iid is not None and ticker:
                _instrument_cache[int(iid)] = str(ticker)
                if item.get("stocksIndustryID") is not None:
                    _industry_cache[str(ticker)] = item["stocksIndustryID"]
        print(f"  {len(_instrument_cache)} instrument i uppslagstabellen.")

    missing = [int(i) for i in instrument_ids if int(i) not in _instrument_cache]
    if missing:
        print(f"    OBS: {len(missing)} instrument kunde inte namnges, behåller ID:n: {missing[:10]}")

    return {int(i): _instrument_cache.get(int(i), str(i)) for i in instrument_ids}


def get_portfolio(username, quiet=False):
    """Hämta en användares live-portfölj.

    Returnerar (weights, meta) där weights = {ticker: vikt%} och
    meta = {ticker: {"öppnad": "ÅÅÅÅ-MM-DD", "vinst_pct": x}} — äldsta öppna
    positionens datum och investeringsviktad vinst. Eller (None, None).
    """
    if not quiet:
        print(f"\nHämtar portfölj för: {username}")
    data = api_get(f"/user-info/people/{username}/portfolio/live", quiet=quiet)
    if not data:
        print(f"  Kunde inte hämta portfölj för '{username}'.")
        return None, None

    # Aggregera per instrumentId (en användare kan ha flera positioner i samma instrument)
    weights = {}     # iid -> summerad investmentPct
    first_open = {}  # iid -> äldsta openTimestamp
    last_open = {}   # iid -> yngsta openTimestamp (senaste köpet)
    profit_acc = {}  # iid -> (summa pct*netProfit, summa pct)

    def add_position(p):
        iid = p.get("instrumentId")
        pct = float(p.get("investmentPct") or 0)
        if iid is None:
            return
        weights[iid] = weights.get(iid, 0) + pct
        ts = p.get("openTimestamp")
        if ts and (iid not in first_open or ts < first_open[iid]):
            first_open[iid] = ts
        if ts and (iid not in last_open or ts > last_open[iid]):
            last_open[iid] = ts
        profit = p.get("netProfit")
        if profit is not None and pct:
            s, w = profit_acc.get(iid, (0.0, 0.0))
            profit_acc[iid] = (s + pct * float(profit), w + pct)

    for p in data.get("positions") or []:
        add_position(p)
    # Positioner inuti social trades (kopierade portföljer) räknas också
    for st in data.get("socialTrades") or []:
        for p in st.get("positions") or []:
            add_position(p)

    if not weights:
        print(f"  Portföljen för '{username}' verkar vara tom eller dold.")
        return None, None

    id_to_ticker = resolve_instruments(list(weights.keys()))
    out = {id_to_ticker[iid]: round(pct, 2) for iid, pct in weights.items()}
    meta = {}
    for iid in weights:
        ticker = id_to_ticker[iid]
        m = {}
        if iid in first_open:
            m["öppnad"] = first_open[iid][:10]
        if iid in last_open:
            m["senaste_köp"] = last_open[iid][:10]
        if iid in profit_acc and profit_acc[iid][1]:
            m["vinst_pct"] = round(profit_acc[iid][0] / profit_acc[iid][1], 1)
        if m:
            meta[ticker] = m
    if not quiet:
        print(f"  {len(out)} innehav hämtade.")
    return out, meta


# ----------------------------------------------------------------------
# Delad konsensusberäkning (används av både signal- och ev. andra grupper)
# ----------------------------------------------------------------------
def konsensus_trosklar(antal_profiler):
    """(in_krav, kvar_krav) i antal ägare för en grupp av given storlek.

    ceil(andel × N) med epsilon-skydd mot flyttalsartefakter (0.6×6 lagras
    som 3.5999... — ceil ger rätt 4, men ett exakt heltal som råkat hamna
    strax ÖVER får inte rundas upp ett extra steg).
    """
    import math
    in_krav = math.ceil(KONSENSUS_ANDEL_IN * antal_profiler - 1e-9)
    kvar_krav = math.ceil(KONSENSUS_ANDEL_KVAR * antal_profiler - 1e-9)
    return in_krav, kvar_krav


def load_previous_consensus():
    """Förra körningens konsensuslista (för hysteresen) ur historikfilen.

    Tom mängd vid kallstart utan historik → IN-tröskeln tillämpas för alla.
    """
    try:
        with open(HISTORY_FILE) as f:
            return set((json.load(f).get("senaste") or {}).get("consensus") or [])
    except (OSError, json.JSONDecodeError):
        return set()


def farskhetsvikt(dagar_sedan_senaste_kop):
    """Hur mycket ett innehav ska väga utifrån hur färskt senaste köpet är.

    Aktivt nyköp (≤30 dgr) väger mer än en gammal vinnare som ligger kvar.
    Saknas datum → neutral 1.0 (degraderar snyggt).

    Caveat: "senaste köp" approximeras med yngsta öppna positionens
    openTimestamp. Partiella stängningar lurar detta — stänger tradern sina
    äldsta lotter och behåller ett gammalt köp ser innehavet ändå ut som det
    yngsta kvarvarande, vilket kan visa högre färskhet än den verkliga
    övertygelsen motiverar.
    """
    if dagar_sedan_senaste_kop is None:
        return 1.0
    if dagar_sedan_senaste_kop <= 30:
        return 1.5   # aktivt köp
    if dagar_sedan_senaste_kop <= 180:
        return 1.0
    return 0.5       # passivt innehav


def compute_consensus(portfolios, port_meta=None, previous_consensus=None):
    """Beräkna konsensus, nära konsensus och bubblarnivå ur portföljerna.

    Procentuella trösklar med HYSTERES: en aktie kommer IN på listan vid
    ceil(KONSENSUS_ANDEL_IN × N) ägare men LIGGER KVAR redan vid
    ceil(KONSENSUS_ANDEL_KVAR × N) — kvarnivån tillämpas bara för aktier
    som var konsensus i föregående körning (previous_consensus).

    Dubbelvillkor (UTBYGGNAD_screener_v2 §1, skalat) gäller ENDAST INnivån:
    antal ägare >= tröskel OCH viktad_konsensus (Σ farskhetsvikt) >= samma
    tal — en ny aktie med gamla/passiva köpare klarar inte innivån bara på
    antal. På KVARnivån (hysteres) styr ENDAST antalsvillkoret listmedlem-
    skapet — en redan etablerad konsensusaktie åker inte ut bara för att
    innehaven åldrats (design­beslut 2026-07-17, TSM-fallet: 3 ägare men
    viktad_konsensus 2,5 < 3,0 skulle annars fällt den ur listan trots
    oförändrat antal ägare). Viktad konsensus påverkar ändå fortsatt
    poängen (Konsensus-komponenten i compute_score_v2), bara inte om
    aktien finns kvar på listan.

    Returnerar (consensus, near_consensus, bubblar_niva):
    - consensus: klarar sin tillämpliga tröskel (in eller kvar)
    - near_consensus: klarar kvarnivåns ANTAL men är inte konsensus
      (ny aktie under innivån, eller fallerat viktkrav)
    - bubblar_niva: exakt en ägare under kvarnivån (bubblar-kandidater)
    Varje entry: {count, avg_weight, total_weight, holders,
    viktad_konsensus, senaste_köp_dagar, tröskel, hysteres}.
    """
    from datetime import date
    port_meta = port_meta or {}
    previous_consensus = previous_consensus or set()
    in_krav, kvar_krav = konsensus_trosklar(len(portfolios))
    today = date.today()
    consensus, near_consensus, bubblar_niva = {}, {}, {}
    all_tickers = set()
    for positions in portfolios.values():
        all_tickers.update(positions.keys())
    for ticker in all_tickers:
        holders = {name: p[ticker] for name, p in portfolios.items() if ticker in p}
        vikter, dagar_lista = [], []
        for name in holders:
            sk = ((port_meta.get(name) or {}).get(ticker) or {}).get("senaste_köp")
            dagar = (today - date.fromisoformat(sk)).days if sk else None
            if dagar is not None:
                dagar_lista.append(dagar)
            vikter.append(farskhetsvikt(dagar))
        hysteres = ticker in previous_consensus
        troskel = kvar_krav if hysteres else in_krav
        entry = {"count": len(holders),
                 "avg_weight": sum(holders.values()) / len(holders),
                 "total_weight": round(sum(holders.values()), 2),
                 "holders": sorted(holders),
                 "viktad_konsensus": round(sum(vikter), 2),
                 "senaste_köp_dagar": min(dagar_lista) if dagar_lista else None,
                 "tröskel": troskel,
                 "hysteres": hysteres}
        if hysteres:
            klarar_konsensus = len(holders) >= troskel   # bara antal på KVAR-nivån
        else:
            klarar_konsensus = len(holders) >= troskel and entry["viktad_konsensus"] >= float(troskel)
        if klarar_konsensus:
            consensus[ticker] = entry
        elif len(holders) >= kvar_krav:
            near_consensus[ticker] = entry
        elif len(holders) == kvar_krav - 1:
            bubblar_niva[ticker] = entry
    return consensus, near_consensus, bubblar_niva


# ----------------------------------------------------------------------
# Steg 0: Screener — bakgrundsgrupp (topp ~50 traders som referens/brusfilter)
#
# Tre separata artefakter:
#   bakgrund_topp50.json — medlemslistan från screenern (--screener).
#     Stabil mellan körningar, kan inspekteras och handjusteras.
#   bakgrund_cache.json  — medlemmarnas portföljer (--divergens hämtar om).
#   Divergensen i default-läget räknas från cachen utan några extra anrop.
# ----------------------------------------------------------------------
BG_MEMBERS_FILE = "bakgrund_topp50.json"
BG_CACHE_FILE = "bakgrund_cache.json"
BG_SIZE = 50
SCREENER_FILTER = {
    "sort": "-gain",                  # minusprefix = fallande
    "pageSize": 100,                  # större pool per period för snittningen
    "gainMin": 15,                    # sållar bort förlorare och nollkonton
    "maxMonthlyRiskScoreMax": 6,      # inga högriskchansare
    "weeksSinceRegistrationMin": 104, # minst 2 år på plattformen
    "copiersMin": 50,                 # sållar BARA bort test-/spökkonton — rankas ej på
    # OBS: isTestAccount och popularInvestor ger 404 — finns ej i denna API-version
}
# API:et saknar TwoYearsAgo (404). "~2 år" approximeras med snittet av två
# perioder: rullande 12 månader + förra kalenderåret — båda måste vara bra.
SCREENER_PERIODS = ("OneYearAgo", "LastYear")


def run_screener():
    """Screena fram bakgrundsgruppen och spara topp 50 till BG_MEMBERS_FILE.

    Rankar på snittgain över SCREENER_PERIODS + låg riskpoäng — INTE på
    antal kopierare. Returnerar listan (eller None vid fel).
    """
    from datetime import date

    pools = {}
    for period in SCREENER_PERIODS:
        print(f"Screenar period {period}...")
        data = api_get("/user-info/people/search", {"period": period, **SCREENER_FILTER})
        if not data or not data.get("items"):
            print(f"  OBS: inga träffar för {period} — avbryter.")
            return None
        pools[period] = {i["userName"]: i for i in data["items"]
                         if i["userName"] not in PROFILES}   # grupperna hålls åtskilda

    # Kräv närvaro i båda perioderna (uthållighet, inte one-hit-wonders)
    gemensamma = set.intersection(*(set(p) for p in pools.values()))
    print(f"  {sum(len(p) for p in pools.values())} kandidater, "
          f"{len(gemensamma)} klarar kraven i båda perioderna.")

    def rank_score(username):
        gains = [pools[p][username]["gain"] for p in SCREENER_PERIODS]
        snitt_gain = sum(gains) / len(gains)
        risk = pools[SCREENER_PERIODS[0]][username]["maxMonthlyRiskScore"]
        return snitt_gain - risk * 5   # riskstraff: en riskpoäng kostar 5 gain-enheter

    rankade = sorted(gemensamma, key=rank_score, reverse=True)[:BG_SIZE]
    medlemmar = []
    for username in rankade:
        i1 = pools[SCREENER_PERIODS[0]][username]
        medlemmar.append({
            "userName": username,
            "gain_1ar": round(i1["gain"], 1),
            "gain_forra_aret": round(pools[SCREENER_PERIODS[1]][username]["gain"], 1),
            "riskpoang": i1["maxMonthlyRiskScore"],
            "veckor_registrerad": i1["weeksSinceRegistration"],
            "copiers": i1["copiers"],
        })

    with open(BG_MEMBERS_FILE, "w") as f:
        json.dump({"datum": date.today().isoformat(),
                   "kriterier": {"perioder": list(SCREENER_PERIODS), **SCREENER_FILTER},
                   "profiler": medlemmar}, f, ensure_ascii=False, indent=2)

    print(f"\nTopp {len(medlemmar)} sparad till {BG_MEMBERS_FILE}:")
    for i, m in enumerate(medlemmar[:10], 1):
        print(f"  {i:>2}. {m['userName']:<22} 1 år {m['gain_1ar']:>6.1f} % | "
              f"förra året {m['gain_forra_aret']:>6.1f} % | risk {m['riskpoang']}")
    if len(medlemmar) > 10:
        print(f"  ... och {len(medlemmar) - 10} till.")
    return medlemmar


def load_background_portfolios(refresh=False):
    """Bakgrundsgruppens portföljer.

    refresh=False (default-läget): läs bara befintlig cache — inga API-anrop.
    refresh=True (--divergens): hämta om portföljerna för medlemmarna i
    BG_MEMBERS_FILE och skriv ny cache.
    """
    import time
    from datetime import date

    cache = {}
    if os.path.exists(BG_CACHE_FILE):
        try:
            with open(BG_CACHE_FILE) as f:
                cache = json.load(f)
        except (json.JSONDecodeError, OSError):
            cache = {}

    if not refresh:
        ports = cache.get("portfolios")
        if ports:
            print(f"\nBakgrundsgrupp: {len(ports)} portföljer från cache ({cache.get('datum', '?')}).")
        return ports

    if not os.path.exists(BG_MEMBERS_FILE):
        raise RuntimeError(f"{BG_MEMBERS_FILE} saknas — kör 'python3 etoro_analys.py --screener' först.")
    with open(BG_MEMBERS_FILE) as f:
        members = json.load(f)
    usernames = [m["userName"] for m in members.get("profiler", [])]
    if not usernames:
        raise RuntimeError(f"{BG_MEMBERS_FILE} innehåller inga profiler.")

    print(f"\nHämtar {len(usernames)} bakgrundsportföljer (screenade {members.get('datum', '?')}, "
          "tar ~1–2 min)...")
    ports = {}
    for username in usernames:
        weights, _ = get_portfolio(username, quiet=True)
        if weights:
            ports[username] = weights
        time.sleep(0.5)   # snäll takt — många anrop i följd

    if not ports:
        print("  OBS: inga portföljer kunde hämtas — behåller gamla cachen.")
        return cache.get("portfolios")

    print(f"  {len(ports)} bakgrundsportföljer hämtade.")
    with open(BG_CACHE_FILE, "w") as f:
        json.dump({"datum": date.today().isoformat(), "medlemslista_datum": members.get("datum"),
                   "portfolios": ports}, f, ensure_ascii=False)
    return ports


def compute_divergence(consensus, portfolios, background):
    """Divergens = andel av signalgruppen som äger aktien − andel av bakgrunden.

    Hög divergens = signalgruppens unika övertygelse (mest intressant).
    Låg/negativ = flockbeteende — 'alla' äger den redan.
    """
    if not background:
        return {}
    bg_n = len(background)
    bg_owners, bg_weightsum = {}, {}
    for weights in background.values():
        for tk, w in weights.items():
            bg_owners[tk] = bg_owners.get(tk, 0) + 1
            bg_weightsum[tk] = bg_weightsum.get(tk, 0) + w

    out = {}
    for tk, info in consensus.items():
        sig = info["count"] / len(portfolios)
        bg = bg_owners.get(tk, 0) / bg_n
        out[tk] = {
            "signal_antal": info["count"],
            "signal_andel_pct": round(sig * 100),
            "bakgrund_antal": bg_owners.get(tk, 0),
            "bakgrund_andel_pct": round(bg * 100, 1),
            "bakgrund_snittvikt": (round(bg_weightsum[tk] / bg_owners[tk], 2)
                                   if bg_owners.get(tk) else 0.0),
            "divergens_pp": round((sig - bg) * 100, 1),
        }
    return out


# ----------------------------------------------------------------------
# Steg 2 & 3: Teknisk analys + analytikerdata (yfinance)
# ----------------------------------------------------------------------
ALPHAVANTAGE_KEY = os.environ.get("ALPHAVANTAGE_API_KEY")
_AV_URL = "https://www.alphavantage.co/query"


def fetch_history_alphavantage(yticker):
    """Reservkälla för dagskurser: Alpha Vantage (Yahoo blockerar ofta molnservrar).

    Returnerar en DataFrame med Open/High/Low/Close/Volume (senaste året) eller None.
    Gratisnyckel: https://www.alphavantage.co/support/#api-key (max 5 anrop/min, 25/dag).
    """
    import time
    import pandas as pd

    if not ALPHAVANTAGE_KEY or "." in yticker:   # bara USA-tickers utan suffix
        return None
    try:
        time.sleep(13)   # gratisnivån tillåter max 5 anrop/minut
        r = requests.get(_AV_URL, timeout=60, params={
            "function": "TIME_SERIES_DAILY", "symbol": yticker,
            "outputsize": "full", "apikey": ALPHAVANTAGE_KEY,
        })
        series = (r.json() or {}).get("Time Series (Daily)")
        if not series:
            return None
        df = pd.DataFrame.from_dict(series, orient="index").astype(float)
        df.index = pd.to_datetime(df.index)
        df = df.sort_index().rename(columns={
            "1. open": "Open", "2. high": "High", "3. low": "Low",
            "4. close": "Close", "5. volume": "Volume"})
        return df[["Open", "High", "Low", "Close", "Volume"]].tail(252)   # ~1 handelsår
    except Exception:
        return None


def fetch_overview_alphavantage(yticker):
    """Analytikerdata från Alpha Vantage OVERVIEW, i samma format som Yahoos info-dict."""
    import time

    if not ALPHAVANTAGE_KEY or "." in yticker:
        return {}
    try:
        time.sleep(13)
        r = requests.get(_AV_URL, timeout=60, params={
            "function": "OVERVIEW", "symbol": yticker, "apikey": ALPHAVANTAGE_KEY,
        })
        d = r.json() or {}

        def num(key, cast=float):
            v = d.get(key)
            try:
                return cast(v)
            except (TypeError, ValueError):
                return None

        counts = {k: num(f"AnalystRating{k}", int) or 0
                  for k in ("StrongBuy", "Buy", "Hold", "Sell", "StrongSell")}
        n = sum(counts.values())
        rec = "n/a"
        if n:
            rec = max(counts, key=counts.get)
            rec = {"StrongBuy": "strong_buy", "Buy": "buy", "Hold": "hold",
                   "Sell": "sell", "StrongSell": "strong_sell"}[rec]
        return {
            "recommendationKey": rec,
            "targetMeanPrice": num("AnalystTargetPrice"),
            "numberOfAnalystOpinions": n or None,
            "sector": d.get("Sector") or None,
            "industry": d.get("Industry") or None,
            "forwardPE": num("ForwardPE"),
            "trailingPE": num("PERatio"),
            "pegRatio": num("PEGRatio"),
            # AV Overview saknar high/low riktkurs och nästa rapportdatum —
            # riktkurs_spridningskvot och nasta_rapport blir None (degraderar snyggt)
        }
    except Exception:
        return {}


def _logga_yf_miss(ticker, falt, exception=None, detalj=None):
    """Diagnostisk logg vid tyst yfinance-fallback (§ felsökning 2026-07-17).
    Skiljer 'fältet fanns inte i svaret' (exception=None, detalj beskriver läget)
    från 'blockering/rate-limit' (exception satt — visar typ + meddelande, t.ex.
    HTTPError 429) — annars ser en tom kolumn likadan ut oavsett orsak.
    """
    if exception is not None:
        print(f"    OBS: yfinance-miss för {ticker}.{falt}: "
              f"{type(exception).__name__}: {exception}")
    else:
        print(f"    OBS: yfinance-miss för {ticker}.{falt}: {detalj or 'tomt/oväntat svar'}")


def next_earnings_date(t, ticker="?"):
    """Nästa kommande rapportdatum (ISO-datum) från yfinance .calendar. None vid miss.

    yfinance har bytt returformat mellan versioner (dict eller DataFrame) —
    hanteras defensivt. Endast Yahoo-källan har detta fält.
    """
    from datetime import date

    try:
        cal = t.calendar
        if cal is None:
            _logga_yf_miss(ticker, "calendar", detalj="calendar var None")
            return None
        if isinstance(cal, dict):
            datum_lista = cal.get("Earnings Date")
        else:
            datum_lista = cal.loc["Earnings Date"].dropna().tolist()
        if not datum_lista:
            _logga_yf_miss(ticker, "calendar", detalj="inget 'Earnings Date' i svaret")
            return None
        if not isinstance(datum_lista, (list, tuple)):
            datum_lista = [datum_lista]
        kandidater = []
        for d in datum_lista:
            try:
                kandidater.append(d if hasattr(d, "year") else date.fromisoformat(str(d)[:10]))
            except Exception:
                continue
        if not kandidater:
            _logga_yf_miss(ticker, "calendar", detalj="kunde inte tolka datumen i svaret")
            return None
        kandidater.sort()
        idag = date.today()
        framtida = [d for d in kandidater if d >= idag]
        return (framtida[0] if framtida else kandidater[-1]).isoformat()
    except Exception as e:
        _logga_yf_miss(ticker, "calendar", e)
        return None


def eps_revision_pct(t, ticker="?"):
    """EPS-estimatrevidering för innevarande räkenskapsår över 90 dgr (%).

    Riktningen på analytikernas vinstestimat — stigande estimat stärker en
    köpsignal, fallande är en klassisk fälla trots hög uppsida. None vid miss.

    Robusthet (prioritetsordning, kvoten exploderar/missvisar annars):
    1. Teckenbyte negativt→positivt: +50 (garanterat över tröskeln, ingen kvot)
    2. Teckenbyte positivt→negativt: -50
    3. abs(90daysAgo) < 0.05: basen för liten för en meningsfull kvot → 0
    4. Annars: kvoten clampad till [-50, 50] innan poängsättning
    """
    try:
        et = t.eps_trend
        if et is None or et.empty or "0y" not in et.index:
            _logga_yf_miss(ticker, "eps_trend", detalj="eps_trend tomt eller saknar '0y'-raden")
            return None
        rad = et.loc["0y"]
        cur, old = float(rad["current"]), float(rad["90daysAgo"])
    except Exception as e:
        _logga_yf_miss(ticker, "eps_trend", e)
        return None
    if old < 0 <= cur:
        return 50.0
    if old > 0 >= cur:
        return -50.0
    if abs(old) < 0.05:
        return 0.0
    pct = (cur - old) / abs(old) * 100
    return round(max(-50.0, min(50.0, pct)), 1)


def analyze_ticker(ticker):
    """Hämta teknisk data + analytikerdata. Yahoo Finance i första hand,
    Stooq som reserv för kursdata (Yahoo blockerar ofta molnservrars IP)."""
    import yfinance as yf
    import pandas as pd

    # eToro-tickers -> Yahoo-tickers (justera vid behov)
    yahoo_map = {"RR.L": "RR.L", "IAG.L": "IAG.L", "KBC.BR": "KBC.BR", "FUR.NV": "FUR.AS", "TI5A.NV": "TI5A.AS"}
    yticker = yahoo_map.get(ticker, ticker)

    hist, info, source = None, {}, "Yahoo"
    try:
        t = yf.Ticker(yticker)
        hist = t.history(period="1y")
        if hist.empty or "Close" not in hist.columns or pd.isna(hist["Close"].iloc[-1]):
            hist = None   # Yahoo kan ge en icke-tom df med NaN-priser när molnservrar blockeras
    except Exception:
        hist = None

    if hist is not None:
        try:
            info = t.info or {}
            if len(info) < 5:
                # Yahoo kan svara 200 med ett nästan tomt info-objekt vid
                # blockering/rate-limit — ingen exception, men lika tyst annars
                _logga_yf_miss(ticker, "info", detalj=f"misstänkt tunt svar ({len(info)} nycklar)")
        except Exception as e:
            _logga_yf_miss(ticker, "info", e)
            info = {}   # kursdata funkade men inte analytikerdatan — fylls från förra körningen
    else:
        hist = fetch_history_alphavantage(yticker)
        source = "Alpha Vantage"
        if hist is not None:
            info = fetch_overview_alphavantage(yticker)

    if hist is None:
        detail = ("varken Yahoo eller Alpha Vantage svarade" if ALPHAVANTAGE_KEY
                  else "Yahoo blockerade och ALPHAVANTAGE_API_KEY saknas")
        return {"ticker": ticker, "error": f"ingen prisdata ({detail})"}

    try:
        close = hist["Close"]
        volume = hist["Volume"]
        price = float(close.iloc[-1])
        ma50 = float(close.rolling(50).mean().iloc[-1])
        ma200_series = close.rolling(200).mean()
        ma200 = float(ma200_series.iloc[-1]) if len(close) >= 200 else None
        # Stiger MA200? Jämför med nivån för ca 1 månad (21 handelsdagar) sedan
        ma200_rising = (float(ma200_series.iloc[-1]) > float(ma200_series.iloc[-21])
                        if len(close) >= 221 else None)

        # RSI(14)
        delta = close.diff()
        gain = delta.clip(lower=0).rolling(14).mean()
        loss = (-delta.clip(upper=0)).rolling(14).mean()
        rs = gain / loss
        rsi = float((100 - 100 / (1 + rs)).iloc[-1])

        # MACD (12, 26, 9)
        ema12 = close.ewm(span=12, adjust=False).mean()
        ema26 = close.ewm(span=26, adjust=False).mean()
        macd_line = ema12 - ema26
        macd_signal = macd_line.ewm(span=9, adjust=False).mean()
        macd_val = float(macd_line.iloc[-1])
        macd_sig = float(macd_signal.iloc[-1])

        # Bollingerband (20, 2) — position i bandet: 0 % = nedre, 100 % = övre
        boll_mid = close.rolling(20).mean()
        boll_std = close.rolling(20).std()
        boll_upper = float((boll_mid + 2 * boll_std).iloc[-1])
        boll_lower = float((boll_mid - 2 * boll_std).iloc[-1])
        boll_pos = ((price - boll_lower) / (boll_upper - boll_lower) * 100
                    if boll_upper > boll_lower else None)

        # 52-veckorsnivåer och momentum
        high52 = float(close.max())
        low52 = float(close.min())
        ret_1m = (price / float(close.iloc[-21]) - 1) * 100 if len(close) >= 21 else None
        ret_3m = (price / float(close.iloc[-63]) - 1) * 100 if len(close) >= 63 else None

        # Volymtrend: snitt senaste 20 dagar mot snitt senaste 3 mån
        vol_20 = float(volume.tail(20).mean())
        vol_90 = float(volume.tail(63).mean())
        vol_trend = (vol_20 / vol_90 - 1) * 100 if vol_90 else None

        rec = info.get("recommendationKey", "n/a")
        target = info.get("targetMeanPrice")
        n_analysts = info.get("numberOfAnalystOpinions")
        upside = round((target / price - 1) * 100, 1) if target else None
        eps_rev = eps_revision_pct(t, ticker) if source == "Yahoo" else None

        # §7 Värdering: forward P/E (fallback trailing), PEG — jämförs mot
        # sektormedian i ett efterföljande steg (build_ranking)
        forward_pe = info.get("forwardPE") or info.get("trailingPE")
        peg_ratio = info.get("pegRatio")

        # §8 Riktkursspridning — hög osäkerhet (>0.8) halverar uppsidepoängen
        target_high, target_low = info.get("targetHighPrice"), info.get("targetLowPrice")
        spridningskvot = None
        if target_high and target_low and target:
            spridningskvot = round((target_high - target_low) / target, 2)

        # §9 Nästa rapportdatum (endast Yahoo — ingen poängeffekt, bara varning)
        nasta_rapport = next_earnings_date(t, ticker) if source == "Yahoo" else None

        # §10 Industry (för sektor→ETF-mappning, halvledare→SOXX)
        industry = info.get("industry")

        # §11 Årlig volatilitet (dagsavkastning stddev × sqrt(252)) för
        # volatilitetsjusterad positionsstorlek — inget poängeffekt
        dagsavkastning = close.pct_change().dropna()
        volatilitet_arlig = (round(float(dagsavkastning.std()) * (252 ** 0.5) * 100, 1)
                             if len(dagsavkastning) >= 20 else None)

        # Kompakt prisserie för candlestick-grafen (senaste ~90 handelsdagarna,
        # med MA50/MA200 som överlägg). Kräver OHLC — finns hos både Yahoo och AV.
        ohlc = []
        if {"Open", "High", "Low"}.issubset(hist.columns):
            ma50_s = close.rolling(50).mean()
            ma200_s = close.rolling(200).mean()

            def _num(v):
                return None if pd.isna(v) else round(float(v), 2)

            tail = hist.tail(90)
            for ts, row in tail.iterrows():
                ohlc.append({
                    "d": ts.strftime("%Y-%m-%d"),
                    "o": _num(row["Open"]), "h": _num(row["High"]),
                    "l": _num(row["Low"]), "c": _num(row["Close"]),
                    "ma50": _num(ma50_s.get(ts)), "ma200": _num(ma200_s.get(ts)),
                })

        return {
            "ticker": ticker,
            "datakälla": source,
            "valuta": info.get("currency") or ("USD" if source == "Alpha Vantage" else None),
            "ohlc": ohlc,
            "pris": round(price, 2),
            "MA50": round(ma50, 2),
            "MA200": round(ma200, 2) if ma200 else None,
            "över_MA200": price > ma200 if ma200 else None,
            "MA200_stigande": ma200_rising,
            "stigande_trend": (price > ma200 and bool(ma200_rising)) if ma200 else None,
            "golden_cross": (ma50 > ma200) if ma200 else None,
            "RSI14": round(rsi, 1),
            "MACD": round(macd_val, 2),
            "MACD_signal": round(macd_sig, 2),
            "MACD_över_signal": macd_val > macd_sig,
            "bollinger_position_%": round(boll_pos, 1) if boll_pos is not None else None,
            "52v_högsta": round(high52, 2),
            "52v_lägsta": round(low52, 2),
            "avstånd_52v_högsta_%": round((price / high52 - 1) * 100, 1),
            "avkastning_1m_%": round(ret_1m, 1) if ret_1m is not None else None,
            "avkastning_3m_%": round(ret_3m, 1) if ret_3m is not None else None,
            "volymtrend_20d_vs_3m_%": round(vol_trend, 1) if vol_trend is not None else None,
            "rekommendation": rec,
            "riktkurs": target,
            "uppsida_%": upside,
            "antal_analytiker": n_analysts,
            "eps_rev_90d_pct": eps_rev,
            "sector": info.get("sector"),   # informativ kolumn — inte klustringsgrund (§4)
            "industry": industry,
            "forward_pe": round(forward_pe, 1) if forward_pe else None,
            "peg_ratio": round(peg_ratio, 2) if peg_ratio else None,
            "riktkurs_hog": target_high, "riktkurs_lag": target_low,
            "riktkurs_spridningskvot": spridningskvot,
            "nasta_rapport": nasta_rapport,
            "volatilitet_arlig_%": volatilitet_arlig,
        }
    except Exception as e:
        return {"ticker": ticker, "error": str(e)}


# ----------------------------------------------------------------------
# Steg 3b: Sammanvägd poäng och rangordning
# ----------------------------------------------------------------------
def compute_correlation_clusters(analyses, tickers, threshold=0.7, min_overlap=40):
    """Korrelationsbaserad klustring av konsensusaktierna (ersätter sektorbaserad).

    Sektoretiketter fångar inte samvariation mellan olika sektorer som ändå
    delar samma tes (t.ex. AMZN/GOOG hör till samma AI-tema som chipklustret,
    fast olika yfinance-sektor). I stället: parvis 63-dagars avkastnings-
    korrelation (priset finns redan i 'ohlc'), greedy single-linkage-kluster
    vid korrelation > threshold. Par med < min_overlap gemensamma dagar
    räknas inte (pandas min_periods → NaN → ingen sammanslagning).

    Returnerar {ticker: {"kluster_id": int, "klusterstorlek": int,
    "klusterfaktor": float}}. Aktier utan (tillräcklig) prisdata hamnar
    ensamma i sitt eget kluster (klusterfaktor 1.0 — neutral degradering).
    """
    import pandas as pd

    serier = {}
    for tk in tickers:
        ohlc = (analyses.get(tk) or {}).get("ohlc")
        if not ohlc:
            continue
        try:
            s = pd.Series([p["c"] for p in ohlc],
                         index=pd.to_datetime([p["d"] for p in ohlc]))
            serier[tk] = s.pct_change().dropna()
        except Exception:
            continue

    parent = {tk: tk for tk in tickers}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    if len(serier) >= 2:
        ret_df = pd.DataFrame(serier)
        corr = ret_df.corr(min_periods=min_overlap)
        cols = list(corr.columns)
        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                val = corr.iloc[i, j]
                if pd.notna(val) and val > threshold:
                    union(cols[i], cols[j])

    grupper = {}
    for tk in tickers:
        grupper.setdefault(find(tk), []).append(tk)
    out = {}
    for idx, (_, medlemmar) in enumerate(sorted(grupper.items()), start=1):
        storlek = len(medlemmar)
        for tk in medlemmar:
            out[tk] = {"kluster_id": idx, "klusterstorlek": storlek,
                       "klusterfaktor": round(1 / (storlek ** 0.5), 3)}
    return out


def compute_netflow_30d(history_log, tickers, dagar=30):
    """Signalgruppens nettoviktändring per aktie senaste `dagar` dagarna (pe).

    Kräver att historiken spänner över >= 7 dagar (>= 2 körningar isär) för
    att vara meningsfull — annars {} (neutral poäng överallt).
    """
    import re
    from datetime import date, timedelta

    if not history_log:
        return {}
    datum = sorted({e["datum"] for e in history_log})
    if len(datum) < 2:
        return {}
    span = (date.fromisoformat(datum[-1]) - date.fromisoformat(datum[0])).days
    if span < 7:
        return {}

    cutoff = (date.today() - timedelta(days=dagar)).isoformat()
    netto = {}
    for e in history_log:
        if e["datum"] < cutoff or e["typ"] not in ("VIKTÄNDRING", "NYTT INNEHAV", "SÅLT INNEHAV"):
            continue
        tal = [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", e["detalj"].replace(",", "."))]
        d = None
        if e["typ"] == "VIKTÄNDRING" and len(tal) >= 2:
            d = tal[1] - tal[0]
        elif e["typ"] == "NYTT INNEHAV" and tal:
            d = tal[0]
        elif e["typ"] == "SÅLT INNEHAV" and tal:
            d = -tal[0]
        if d is not None and e["ticker"] in tickers:
            netto[e["ticker"]] = netto.get(e["ticker"], 0.0) + d
    return {tk: round(v, 2) for tk, v in netto.items()}


def compute_sector_pe_medians(analyses, tickers):
    """Grov forward P/E-median per sektor, beräknad över analyserade aktier
    i denna körning (konsensus). Litet urval — 'grov' är avsiktligt."""
    from statistics import median
    grupp = {}
    for tk in tickers:
        a = analyses.get(tk) or {}
        sektor, fpe = a.get("sector"), a.get("forward_pe")
        if sektor and fpe and fpe > 0:
            grupp.setdefault(sektor, []).append(fpe)
    return {sektor: median(vals) for sektor, vals in grupp.items()}


def compute_valuation_score(a, sector_medians):
    """§7 Värderingspoäng (0–10): forward P/E mot sektormedian + PEG-bonus.

    Saknas BÅDA underlagen (ingen P/E-jämförelse OCH ingen PEG) → neutral
    5.0 (degraderingsprincipen: saknad data ska aldrig straffa som "dyr
    aktie", 0 var tidigare omöjligt att skilja från genuint neutral
    värdering). Finns minst ett av underlagen räknas poängen som vanligt.
    Returnerar (poäng, är_neutral).
    """
    fpe, peg = a.get("forward_pe"), a.get("peg_ratio")
    median_val = sector_medians.get(a.get("sector"))
    pe_kand = bool(fpe and median_val)
    peg_kand = peg is not None
    if not pe_kand and not peg_kand:
        return 5.0, True

    varde = 0.0
    if pe_kand:
        if fpe < 0.8 * median_val:
            varde += 10
        elif fpe > 1.5 * median_val:
            varde -= 10
    if peg_kand and 0 < peg < 1.5:
        varde += 5
    return round(max(0.0, min(10.0, varde)), 1), False


# §10 Relativ styrka mot sektor — ETF-mappning (halvledare→SOXX oavsett sektor)
_SEKTOR_ETF = {
    "Technology": "XLK", "Consumer Cyclical": "XLY", "Communication Services": "XLC",
    "Financial Services": "XLF", "Financial": "XLF", "Healthcare": "XLV",
    "Industrials": "XLI", "Energy": "XLE",
}


def _etf_for(sector, industry):
    if industry and "Semiconductor" in industry:
        return "SOXX"
    return _SEKTOR_ETF.get(sector, "SPY")


def _fetch_etf_return_63d(cache, etf_ticker):
    """63-dagarsavkastning för ett index-ETF, cachead per körning (max ~9 ETF:er)."""
    if etf_ticker in cache:
        return cache[etf_ticker]
    try:
        import yfinance as yf
        h = yf.Ticker(etf_ticker).history(period="6mo")
        if h.empty or len(h) < 64:
            cache[etf_ticker] = None
        else:
            close = h["Close"]
            cache[etf_ticker] = round((float(close.iloc[-1]) / float(close.iloc[-63]) - 1) * 100, 1)
    except Exception:
        cache[etf_ticker] = None
    return cache[etf_ticker]


def compute_relative_strength(analyses, tickers):
    """§10: RS = aktiens 63d-avkastning − dess sektor-ETF:s 63d-avkastning.

    Returnerar {ticker: {"rs_pe": x, "etf": "XLK", "bonus": ±5}}. ETF-priser
    hämtas högst en gång per unikt index i hela körningen (yfinance, ingen
    AV-reserv här — degraderar till None om Yahoo blockerar).
    """
    cache = {}
    out = {}
    for tk in tickers:
        a = analyses.get(tk) or {}
        r3 = a.get("avkastning_3m_%")
        etf = _etf_for(a.get("sector"), a.get("industry"))
        etf_ret = _fetch_etf_return_63d(cache, etf)
        if r3 is None or etf_ret is None:
            out[tk] = {"rs_pe": None, "etf": etf, "bonus": 0}
            continue
        rs = round(r3 - etf_ret, 1)
        out[tk] = {"rs_pe": rs, "etf": etf,
                  "bonus": 5 if rs > 5 else (-5 if rs < -5 else 0)}
    return out


def compute_suggested_weights(analyses, tickers, malvikt=2.0, tak=3.0, golv=0.5):
    """§11: Volatilitetsjusterad föreslagen vikt (%) — sizing, ingen poängeffekt.

    raw = 1/volatilitet, normaliserad så snittförslaget blir `malvikt` %,
    därefter clampad till [golv, tak]. Ingen poängeffekt.
    """
    raw = {}
    for tk in tickers:
        v = (analyses.get(tk) or {}).get("volatilitet_arlig_%")
        if v and v > 0:
            raw[tk] = 100.0 / v
    if not raw:
        return {}
    medel = sum(raw.values()) / len(raw)
    if not medel:
        return {}
    skala = malvikt / medel
    return {tk: round(min(tak, max(golv, r * skala)), 2) for tk, r in raw.items()}


def compute_score(a, cons, cluster_factor=1.0, nettoflode_pe=None):
    """Sammanvägd poäng 0–100 för en aktie.

    Trend 30 p + Momentum 25 p + Analytiker 25 p + Konsensus 20 p (justerad
    för korrelationskluster §4 och nettoflöde §5). Returnerar (totalpoäng,
    delpoäng-dict).
    """
    delpoang = {}

    # Trend (max 30) — investerarens viktigaste kriterium
    trend = 0
    if a.get("över_MA200"):
        trend += 10
    if a.get("MA200_stigande"):
        trend += 10
    if a.get("golden_cross"):
        trend += 10
    delpoang["Trend"] = trend

    # Momentum (max 25)
    mom = 0.0
    rsi = a.get("RSI14")
    if rsi is not None:
        if 45 <= rsi <= 65:      # styrkezon utan överköp
            mom += 10
        elif 35 <= rsi < 45 or 65 < rsi <= 70:
            mom += 6
        elif 30 <= rsi < 35:
            mom += 3
    if a.get("MACD_över_signal"):
        mom += 7
    r3 = a.get("avkastning_3m_%")
    if r3 is not None:
        mom += max(0.0, min(8.0, r3 / 5))   # +40 % på 3 mån ger full poäng
    delpoang["Momentum"] = round(mom, 1)

    # Analytiker (max 25): uppsida + antal + köprek + estimatrevidering
    ana = 0.0
    uppsida = a.get("uppsida_%")
    if uppsida is not None:
        ana += max(0.0, min(15.0, uppsida / 2))   # 30 % uppsida ger full poäng
    ana += min(5.0, (a.get("antal_analytiker") or 0) / 8)   # 40 analytiker ger full poäng
    if a.get("rekommendation") in ("strong_buy", "buy"):
        ana += 5
    eps_rev = a.get("eps_rev_90d_pct")
    if eps_rev is not None:   # stigande estimat +, fallande − (fångar uppsidefällan)
        ana += 10 if eps_rev > 5 else (-10 if eps_rev < -5 else 0)
    delpoang["Analytiker"] = round(max(0.0, min(25.0, ana)), 1)

    # Konsensus (max 20, före klusterjustering) — eniga, övertygade OCH
    # färska investerare, plus flödesriktning; delat med sqrt(klusterstorlek)
    # om aktien samvarierar starkt med andra konsensusaktier (§4)
    kon = 0.0
    vk = cons.get("viktad_konsensus")
    if vk is None:
        vk = float(cons["count"])   # degradera snyggt om färskhetsdata saknas
    kon += min(12.0, vk * 3.0)      # vk=3 (tröskel) → 9; vk≥4 → 12; stale trio (1.5) → 4.5
    kon += min(8.0, cons["avg_weight"] * 1.6)   # ~5 % snittvikt ger full poäng
    if nettoflode_pe is not None:
        kon += 5 if nettoflode_pe > 1.0 else (-5 if nettoflode_pe < -1.0 else 0)
    kon *= cluster_factor   # ensam i sitt kluster → oförändrad (faktor 1.0)
    delpoang["Konsensus"] = round(max(0.0, kon), 1)

    total = round(sum(delpoang.values()), 1)
    return total, delpoang


def compute_score_v2(a, cons, cluster_factor=1.0, nettoflode_pe=None,
                     rs_bonus=0, sector_medians=None):
    """§12 Omviktad poängmodell: Trend 25 + Momentum 20 + Analytiker 20 +
    Konsensus 25 + Värdering 10 (vikterna är startvärden — --utvardera ska
    på sikt kalibrera dem). Återanvänder v1:s delformler och klipper varje
    komponent till sin nya takhöjd, plus: relativ styrka (§10) i Momentum,
    riktkursspridning (§8) halverar uppsidan i Analytiker, Värdering (§7)
    är helt ny. Returnerar (totalpoäng, delpoäng-dict).
    """
    sector_medians = sector_medians or {}
    delpoang = {}

    # Trend (tak 25, samma flaggor som v1 vars råsumma är 0..30)
    trend_raw = 0
    if a.get("över_MA200"):
        trend_raw += 10
    if a.get("MA200_stigande"):
        trend_raw += 10
    if a.get("golden_cross"):
        trend_raw += 10
    delpoang["Trend"] = min(25, trend_raw)

    # Momentum (tak 20): v1:s bas (0..25) + relativ styrka mot sektor-ETF (±5)
    mom = 0.0
    rsi = a.get("RSI14")
    if rsi is not None:
        if 45 <= rsi <= 65:
            mom += 10
        elif 35 <= rsi < 45 or 65 < rsi <= 70:
            mom += 6
        elif 30 <= rsi < 35:
            mom += 3
    if a.get("MACD_över_signal"):
        mom += 7
    r3 = a.get("avkastning_3m_%")
    if r3 is not None:
        mom += max(0.0, min(8.0, r3 / 5))
    mom += rs_bonus
    delpoang["Momentum"] = round(max(0.0, min(20.0, mom)), 1)

    # Analytiker (tak 20): uppsidan halveras vid hög riktkursspridning (§8)
    ana = 0.0
    uppsida = a.get("uppsida_%")
    if uppsida is not None:
        uppsida_poang = max(0.0, min(15.0, uppsida / 2))
        if (a.get("riktkurs_spridningskvot") or 0) > 0.8:
            uppsida_poang /= 2
        ana += uppsida_poang
    ana += min(5.0, (a.get("antal_analytiker") or 0) / 8)
    if a.get("rekommendation") in ("strong_buy", "buy"):
        ana += 5
    eps_rev = a.get("eps_rev_90d_pct")
    if eps_rev is not None:
        ana += 10 if eps_rev > 5 else (-10 if eps_rev < -5 else 0)
    delpoang["Analytiker"] = round(max(0.0, min(20.0, ana)), 1)

    # Konsensus (tak 25) — identisk innehåll som v1, nya taket råkar matcha
    # exakt (12+8+5=25) så ingen omskalning behövs
    kon = 0.0
    vk = cons.get("viktad_konsensus")
    if vk is None:
        vk = float(cons["count"])
    kon += min(12.0, vk * 3.0)
    kon += min(8.0, cons["avg_weight"] * 1.6)
    if nettoflode_pe is not None:
        kon += 5 if nettoflode_pe > 1.0 else (-5 if nettoflode_pe < -1.0 else 0)
    kon *= cluster_factor
    delpoang["Konsensus"] = round(max(0.0, min(25.0, kon)), 1)

    # Värdering (tak 10) — helt ny komponent (§7)
    värdering_poäng, värdering_neutral = compute_valuation_score(a, sector_medians)
    delpoang["Värdering"] = värdering_poäng

    total = round(sum(delpoang.values()), 1)
    return total, delpoang, värdering_neutral


# ----------------------------------------------------------------------
# §A Marknadsregimfilter (UTBYGGNAD_regim_exit.md)
# ----------------------------------------------------------------------
REGIM_TICKER_KEDJA = ["SPY", "^GSPC", "VOO", "IVV"]   # alla S&P 500 — identisk MA200-regim


def _hamta_regim_for_ticker(ticker):
    """Ett försök i regimkedjan. Returnerar {regim, spy_pris, spy_ma200} vid
    lyckad beräkning, annars None (och loggar orsaken via _logga_yf_miss)."""
    import yfinance as yf
    try:
        h = yf.Ticker(ticker).history(period="1y", auto_adjust=True)
        if h.empty or len(h) < 221:
            _logga_yf_miss(ticker, "history",
                          detalj=f"otillräcklig historik ({len(h)} rader, behöver ≥221)")
            return None
        close = h["Close"]
        pris = float(close.iloc[-1])
        ma200_serie = close.rolling(200).mean()
        ma200 = float(ma200_serie.iloc[-1])
        ma200_stigande = ma200 > float(ma200_serie.iloc[-21])
        over = pris > ma200
        if over and ma200_stigande:
            regim = "GRÖN"
        elif not over and not ma200_stigande:
            regim = "RÖD"
        else:
            regim = "GUL"
        return {"regim": regim, "spy_pris": round(pris, 2), "spy_ma200": round(ma200, 2)}
    except Exception as e:
        _logga_yf_miss(ticker, "history", e)
        return None


def compute_market_regime():
    """Marknadsregim (index vs MA200): GRÖN (över + stigande), RÖD (under
    + fallande), annars GUL (blandat, bara varning).

    Provar REGIM_TICKER_KEDJA (SPY, ^GSPC, VOO, IVV — alla S&P 500-trackare,
    ger identisk MA200-regim) i tur och ordning tills en lyckas, så att en
    enskild blockerad/rate-limitad ticker (t.ex. SPY specifikt på Render,
    se CLAUDE.md § Kända miljöbegränsningar) inte slår ut regimberäkningen
    helt. `regim_kalla` anger vilken ticker som faktiskt lyckades.

    Misslyckas ALLA fyra → OKÄND (regim_kalla=None). Nedströms (run_analysis)
    behandlas OKÄND identiskt med GRÖN och en senaste kända regim återanvänds
    om en finns (hellre falskt grönt/inaktuellt än att blockera på datafel).
    """
    from datetime import date
    for ticker in REGIM_TICKER_KEDJA:
        resultat = _hamta_regim_for_ticker(ticker)
        if resultat is not None:
            resultat["regim_kalla"] = ticker
            resultat["notis"] = None
            resultat["datum"] = date.today().isoformat()
            resultat["regim_datum"] = date.today().isoformat()
            return resultat
    return {"regim": "OKÄND", "spy_pris": None, "spy_ma200": None, "regim_kalla": None,
            "notis": f"alla index i fallbackkedjan misslyckades ({', '.join(REGIM_TICKER_KEDJA)})",
            "datum": date.today().isoformat(), "regim_datum": None}


REGIM_ALDER_VARNING_HANDELSDAGAR = 5


def _handelsdagar_mellan(datum1_str, datum2_str):
    """Grov approximation av antal handelsdagar (vardagar mån–fre) mellan två
    ISO-datum, utan hänsyn till marknadshelgdagar — tillräckligt för en
    varningströskel, ingen exakt handelskalender behövs."""
    from datetime import date, timedelta
    d1, d2 = date.fromisoformat(datum1_str), date.fromisoformat(datum2_str)
    if d2 < d1:
        d1, d2 = d2, d1
    dagar, d = 0, d1
    while d < d2:
        d += timedelta(days=1)
        if d.weekday() < 5:
            dagar += 1
    return dagar


# ----------------------------------------------------------------------
# §B Exitregel (trendbrott) (UTBYGGNAD_regim_exit.md)
# ----------------------------------------------------------------------
EXIT_VILLKOR_TEXT = "Dödskors: pris < MA200 och MA50 < MA200"


def compute_exit_status(analyses, tickers):
    """Exitvillkor: pris < MA200 OCH MA50 < MA200 (båda krävs — pris under
    MA200 ensamt är en normal rekyl). Saknas MA200/MA50 (kort historik) →
    False (kan inte avgöras, aldrig krasch)."""
    out = {}
    for tk in tickers:
        a = analyses.get(tk) or {}
        pris, ma50, ma200 = a.get("pris"), a.get("MA50"), a.get("MA200")
        out[tk] = bool(pris is not None and ma50 is not None and ma200 is not None
                       and pris < ma200 and ma50 < ma200)
    return out


def build_ranking(analyses, consensus, history_log=None, exit_info=None):
    """Rangordna konsensusaktierna. Aktier utan stigande trend sist, oavsett poäng.

    Primär poäng = §12:s omviktade modell (Trend25/Momentum20/Analytiker20/
    Konsensus25/Värdering10). Gamla v1-poängen (Trend30/Momentum25/
    Analytiker25/Konsensus20) sparas som poäng_v1/delpoäng_v1 för jämförelse
    under övergångsperioden, tills --utvardera hunnit kalibrera vikterna.

    exit_info: {ticker: exit_datum} för aktier i EXIT (§B trendbrott) — de
    beräknas som vanligt (poängen får inte förorenas) men returneras separat
    i exit_lista i stället för ranking, så Bästa köp/pappersportföljerna
    utesluter dem automatiskt. Konsensuslistan påverkas inte av detta.
    """
    exit_info = exit_info or {}
    giltiga = [tk for tk in consensus if "error" not in analyses.get(tk, {})]
    kluster = compute_correlation_clusters(analyses, giltiga)
    netto = compute_netflow_30d(history_log or [], giltiga)
    sector_medians = compute_sector_pe_medians(analyses, giltiga)
    rs = compute_relative_strength(analyses, giltiga)
    forslagen_vikt = compute_suggested_weights(analyses, giltiga)

    ranking, exit_lista = [], []
    for ticker, cons in consensus.items():
        a = analyses.get(ticker, {})
        if "error" in a:
            continue
        kf = kluster.get(ticker, {}).get("klusterfaktor", 1.0)
        nf = netto.get(ticker)
        rs_bonus = rs.get(ticker, {}).get("bonus", 0)

        total, delpoang, värdering_neutral = compute_score_v2(
            a, cons, cluster_factor=kf, nettoflode_pe=nf,
            rs_bonus=rs_bonus, sector_medians=sector_medians)
        total_v1, delpoang_v1 = compute_score(a, cons, cluster_factor=kf, nettoflode_pe=nf)

        rad = {
            "ticker": ticker,
            "poäng": total,
            "poäng_v1": total_v1,
            "trend_ok": bool(a.get("stigande_trend")),
            "delpoäng": delpoang,
            "delpoäng_v1": delpoang_v1,
            "vardering_neutral": värdering_neutral,
            "kluster": kluster.get(ticker),
            "nettoflode_30d_pe": nf,
            "relativ_styrka": rs.get(ticker),
            "foreslagen_vikt_%": forslagen_vikt.get(ticker),
        }
        if ticker in exit_info:
            rad["exit_datum"] = exit_info[ticker]
            rad["exit_villkor"] = EXIT_VILLKOR_TEXT
            exit_lista.append(rad)
        else:
            ranking.append(rad)
    ranking.sort(key=lambda r: (not r["trend_ok"], -r["poäng"]))
    exit_lista.sort(key=lambda r: r["exit_datum"])
    return ranking, exit_lista


# ----------------------------------------------------------------------
# Steg 4 — Claude-triggerfilter: bara väsentliga förändringar omanalyseras
# ----------------------------------------------------------------------
CLAUDE_MODELL_NY = "claude-opus-4-8"        # grundanalys — aktien saknar text
CLAUDE_MODELL_OMANALYS = "claude-sonnet-4-6"   # omanalys av befintlig aktie
MODELL_KORTNAMN = {CLAUDE_MODELL_NY: "opus", CLAUDE_MODELL_OMANALYS: "sonnet"}
MAX_ANALYS_ALDER_DAGAR = 7
POANG_TRIGGER_DIFF = 10.0
KONSENSUS_TRIGGER_DIFF = 1.0
# Output-tak: Sonnet-omanalyser kör UTAN thinking (ren textbudget, ~120 ord
# räcker med marginal). Opus-grundanalyser behåller adaptive thinking, som
# äter av samma max_tokens-budget — mer marginal krävs där (se incidenten
# 2026-07-16 där 2000 utan uppdelning gav tomma svar för två aktier).
CLAUDE_MAX_TOKENS_OMANALYS = 600
CLAUDE_MAX_TOKENS_NY = 2000


def _bygg_indikator_snapshot(a, poang, viktad_konsensus, exit_flagga, regim_status=None):
    """Kompakt ögonblicksbild av det som avgör om en Claude-text blivit inaktuell.

    Sparas bredvid varje analystext (claude[tk].indikator_snapshot) så nästa
    körning kan jämföra dagens värden mot de som gällde när texten skrevs.
    regim_status sparas ENDAST för triggerjämförelsen (GRÖN↔RÖD-skifte) —
    regimen skickas inte längre till Claude i själva prompten (se
    _bygg_claude_input), den visas redan i rapportheadern.
    """
    macd, macd_sig = a.get("MACD"), a.get("MACD_signal")
    return {
        "rsi": a.get("RSI14"),
        "pris": a.get("pris"),
        "ma50": a.get("MA50"),
        "ma200": a.get("MA200"),
        "macd_diff": round(macd - macd_sig, 4) if macd is not None and macd_sig is not None else None,
        "over_ma200": a.get("över_MA200"),
        "golden_cross": a.get("golden_cross"),
        "poang": poang,
        "viktad_konsensus": viktad_konsensus,
        "exit": exit_flagga,
        "regim": regim_status,
    }


def _rsi_korsade_troskel(old_rsi, new_rsi, troskel):
    if old_rsi is None or new_rsi is None:
        return False
    return (old_rsi < troskel <= new_rsi) or (new_rsi < troskel <= old_rsi)


def behover_ny_analys(ticker, dagens_data, senaste_analys):
    """Avgör om `ticker` behöver Claude-omanalyseras. Returnerar (bool, orsak).

    dagens_data: dagens indikator_snapshot (se _bygg_indikator_snapshot).
    senaste_analys: förra körningens claude[ticker]-dict, eller None/tomt om
    aktien saknar sparad text (ny på listan/kallstart) — analyseras då alltid.
    Saknar den befintliga texten fältet indikator_snapshot (gammal fil från
    före detta filter) behandlas den likaså som saknad — engångsomanalys,
    därefter normal jämförelse.

    Analyseras om ENDAST vid: saknad text, RSI-korsning av 30/70, pris-korsning
    av MA200, MACD-korsning av signallinjen, golden/death cross, ändrad
    EXIT-status, poängändring > 10, viktad konsensus-ändring > 1.0,
    regimskifte GRÖN↔RÖD (GUL/OKÄND triggar inte — för brusigt), eller
    text äldre än MAX_ANALYS_ALDER_DAGAR dagar.
    """
    from datetime import date
    if not senaste_analys or not senaste_analys.get("analys"):
        return True, "ny på listan"

    gammal = senaste_analys.get("indikator_snapshot")
    if not gammal:
        return True, "saknar indikatorsnapshot (gammal analys, engångsuppdatering)"

    genererad = senaste_analys.get("genererad")
    if genererad:
        alder = (date.today() - date.fromisoformat(genererad)).days
        if alder > MAX_ANALYS_ALDER_DAGAR:
            return True, f"text äldre än {MAX_ANALYS_ALDER_DAGAR} dagar ({alder} dagar)"

    if _rsi_korsade_troskel(gammal.get("rsi"), dagens_data.get("rsi"), 30):
        return True, "RSI korsade 30"
    if _rsi_korsade_troskel(gammal.get("rsi"), dagens_data.get("rsi"), 70):
        return True, "RSI korsade 70"

    go, gn = gammal.get("over_ma200"), dagens_data.get("over_ma200")
    if go is not None and gn is not None and go != gn:
        return True, "pris korsade MA200"

    gmd, nmd = gammal.get("macd_diff"), dagens_data.get("macd_diff")
    if gmd is not None and nmd is not None and (gmd > 0) != (nmd > 0):
        return True, "MACD korsade signallinjen"

    ggc, ngc = gammal.get("golden_cross"), dagens_data.get("golden_cross")
    if ggc is not None and ngc is not None and ggc != ngc:
        return True, "dödskors/goldenkors inträffade"

    ge, ne = gammal.get("exit"), dagens_data.get("exit")
    if ge is not None and ne is not None and ge != ne:
        return True, ("aktien gick in i EXIT" if ne else "aktien återinträdde från EXIT")

    gp, np_ = gammal.get("poang"), dagens_data.get("poang")
    if gp is not None and np_ is not None and abs(np_ - gp) > POANG_TRIGGER_DIFF:
        return True, f"poäng ändrad {gp:.1f} → {np_:.1f}"

    gk, nk = gammal.get("viktad_konsensus"), dagens_data.get("viktad_konsensus")
    if gk is not None and nk is not None and abs(nk - gk) > KONSENSUS_TRIGGER_DIFF:
        return True, f"viktad konsensus ändrad {gk:.1f} → {nk:.1f}"

    gr, nr = gammal.get("regim"), dagens_data.get("regim")
    if {gr, nr} == {"GRÖN", "RÖD"}:   # GUL/OKÄND ingår aldrig i mängden -> triggar inte
        return True, f"regimskifte {gr} → {nr}"

    return False, "inga väsentliga ändringar"


def _bygg_claude_input(ticker, a, poäng, delpoäng, viktad_konsensus, divergens_pp,
                       exit_flagga):
    """Tokensnål payload till Claude — bara fälten systemprompten faktiskt
    använder, ALDRIG hela indikator-dicten eller rå yfinance-data. Flyttal
    avrundas till 1–2 decimaler (långa decimaler kostar tokens utan att
    tillföra något); nyckar med None-värde utesluts helt.

    Regim skickas INTE med (visas redan i rapportheadern, och en återanvänd
    text skulle annars bära en inaktuell regim-etikett) — regimskifte
    GRÖN↔RÖD är i stället en omanalys-trigger, se behover_ny_analys.
    """
    def r(x, n=1):
        return round(x, n) if isinstance(x, (int, float)) else x

    payload = {
        "ticker": ticker,
        "pris": r(a.get("pris"), 2),
        "valuta": a.get("valuta"),
        "RSI14": r(a.get("RSI14")),
        "MA50": r(a.get("MA50"), 2),
        "MA200": r(a.get("MA200"), 2),
        "stigande_trend": a.get("stigande_trend"),
        "MACD_över_signal": a.get("MACD_över_signal"),
        "golden_cross": a.get("golden_cross"),
        "bollinger_position_%": r(a.get("bollinger_position_%")),
        "avstånd_52v_högsta_%": r(a.get("avstånd_52v_högsta_%")),
        "avkastning_1m_%": r(a.get("avkastning_1m_%")),
        "avkastning_3m_%": r(a.get("avkastning_3m_%")),
        "poäng": r(poäng),
        "delpoäng": {k: r(v) for k, v in (delpoäng or {}).items()} or None,
        "viktad_konsensus": r(viktad_konsensus),
        "divergens_mot_bakgrundsgruppen_pp": r(divergens_pp),
        "eps_rev_90d_pct": r(a.get("eps_rev_90d_pct")),
        "riktkurs": r(a.get("riktkurs"), 2),
        "riktkurs_spridningskvot": r(a.get("riktkurs_spridningskvot"), 2),
        "uppsida_%": r(a.get("uppsida_%")),
        "investerarnas_innehavstid_dagar_längst": a.get("investerarnas_innehavstid_dagar_längst"),
        "investerarnas_innehavstid_dagar_snitt": a.get("investerarnas_innehavstid_dagar_snitt"),
        "investerarnas_upparbetade_vinst_pct_snitt": r(a.get("investerarnas_upparbetade_vinst_pct_snitt")),
        "nasta_rapport": a.get("nasta_rapport"),
        "exit": exit_flagga,
    }
    return {k: v for k, v in payload.items() if v is not None}


# ----------------------------------------------------------------------
# Steg 4: Claude gör en gedigen teknisk analys per konsensusaktie
# ----------------------------------------------------------------------
def claude_analysis(jobb, körningsläge="standard"):
    """Skicka indikatordata till Claude API och få en skriven analys per aktie.

    jobb: {ticker: {"data": analysdikt, "model": modell-id, "orsak": str,
    "snapshot": indikator_snapshot}} — modell och orsak väljs av anroparen
    (§ Claude-triggerfilter): claude-opus-4-8 för "ny på listan"
    (grundanalys), annars claude-sonnet-4-6 (omanalys av befintlig aktie).
    körningsläge: "standard" | "divergens" | "force" — taggas på varje
    tokenförbrukningspost (se logga_forbrukning).

    Returnerar (results, forbrukning):
    - results: {ticker: {"rekommendation", "analys", "genererad", "modell",
      "analys_orsak", "indikator_snapshot"}}.
    - forbrukning: lista med en tokenförbrukningspost per lyckat anrop
      (se logga_forbrukning för fältformat). Saknas usage-objektet i
      API-svaret (äldre SDK e.d.) loggas posten ändå, med tokenfälten null
      och en varning i terminalen — kraschar aldrig.

    Hoppar över (med varning) om ANTHROPIC_API_KEY saknas.
    """
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("\nOBS: ANTHROPIC_API_KEY saknas i miljön/.env — hoppar över Claude-analysen.")
        return {}, []

    import anthropic
    client = anthropic.Anthropic()

    system = (
        "Du är en erfaren teknisk analytiker på en svensk bank. Du får tekniska "
        "indikatorer och analytikerdata för en aktie i JSON-format (fälten är redan "
        "urvalda — inga andra data finns). Skriv en koncis teknisk analys på svenska "
        "(max ca 120 ord) som väger samman trend (MA50/MA200, golden cross), momentum "
        "(RSI, MACD, 1m/3m-avkastning), volatilitet (Bollingerband, avstånd till "
        "52-veckorsnivåer) och analytikernas riktkurs. Var konkret: nämn nivåer och "
        "vad som skulle ändra bilden.\n\n"
        "VIKTIGASTE KRITERIET för investeraren är stigande trend (fältet "
        "stigande_trend). Är den false får rekommendationen ALDRIG vara KÖP — högst "
        "AVVAKTA, och ange då tydligt vilken nivå som måste återtas för att "
        "trendkriteriet ska vara uppfyllt. Är trendkriteriet uppfyllt, bedöm övriga "
        "indikatorer som vanligt.\n\n"
        "VINSTHEMTAGNINGSRISK: fälten investerarnas_innehavstid_dagar_* och "
        "investerarnas_upparbetade_vinst_pct_snitt visar hur länge eToro-investerarna "
        "ägt aktien och deras upparbetade vinst. Lång innehavstid i kombination med "
        "hög upparbetad vinst ökar risken att de börjar sälja och ta hem vinsten — "
        "väg in det i bedömningen och kommentera det uttryckligen när risken är "
        "förhöjd.\n\n"
        "DIVERGENS: fältet divergens_mot_bakgrundsgruppen_pp visar hur mycket mer "
        "(eller mindre) signalgruppen äger aktien jämfört med en bred referensgrupp "
        "(topp ~50 screenade traders). Hög divergens = unik övertygelse (starkare "
        "signal); låg/negativ = flockbeteende. Nämn det kort.\n\n"
        "EXIT: fältet exit=true betyder att aktien har ett dödskors (pris och MA50 "
        "båda under MA200) — betona att trenden brutits även om andra indikatorer "
        "ser bra ut.\n\n"
        "VALUTA: alla priser och nivåer anges i aktiens handelsvaluta (fältet "
        "'valuta', t.ex. USD, EUR, SEK). Använd rätt valutabeteckning — skriv "
        "ALDRIG 'kr' på ett USD-pris. Är valutan USD, skriv priser som t.ex. "
        "'242 USD' eller '$242', aldrig kronor.\n\n"
        "SPRÅK: skriv naturligt löpande språk, som till en investerare — använd "
        "ALDRIG råa fältnamn i texten (t.ex. 'stigande_trend är false' eller "
        "'exit=true'), skriv i stället vad det betyder i klartext. Använd heller "
        "ingen markdown-formatering (inga **, #, listor) — ren löptext.\n\n"
        "Svara EXAKT i detta format:\n"
        "REKOMMENDATION: <KÖP | AVVAKTA | SÄLJ>\n"
        "<själva analysen>"
    )

    from datetime import date, datetime
    idag = date.today().isoformat()
    results = {}
    forbrukning = []
    for ticker, job in jobb.items():
        data = job["data"]
        modell = job.get("model") or CLAUDE_MODELL_NY
        orsak = job.get("orsak", "?")
        # Sonnet-omanalyser: ingen thinking (ren textbudget). Opus-grundanalyser
        # ("ny på listan"): behåller adaptive thinking, med mer max_tokens-marginal.
        är_ny = modell == CLAUDE_MODELL_NY
        thinking_param = {"type": "adaptive"} if är_ny else {"type": "disabled"}
        max_tok = CLAUDE_MAX_TOKENS_NY if är_ny else CLAUDE_MAX_TOKENS_OMANALYS
        print(f"  Claude analyserar {ticker} ({modell}, orsak: {orsak})...")
        try:
            resp = client.messages.create(
                model=modell,
                max_tokens=max_tok,
                thinking=thinking_param,
                system=system,
                messages=[{
                    "role": "user",
                    "content": f"Analysera {ticker}:\n{json.dumps(data, ensure_ascii=False, indent=2)}",
                }],
            )
            text = next((b.text for b in resp.content if b.type == "text"), "").strip()
            rating = "?"
            if text.upper().startswith("REKOMMENDATION:"):
                first, _, rest = text.partition("\n")
                rating = first.split(":", 1)[1].strip()
                text = rest.strip()
            results[ticker] = {
                "rekommendation": rating, "analys": text, "genererad": idag,
                "modell": modell, "analys_orsak": orsak,
                "indikator_snapshot": job.get("snapshot"),
            }

            usage = getattr(resp, "usage", None)
            if usage is None:
                print(f"    OBS: usage-data saknas i Claude-svaret för {ticker} (äldre SDK-version?).")
                in_tok = out_tok = cache_skapad = cache_last = None
            else:
                in_tok = getattr(usage, "input_tokens", None)
                out_tok = getattr(usage, "output_tokens", None)
                cache_skapad = getattr(usage, "cache_creation_input_tokens", None)
                cache_last = getattr(usage, "cache_read_input_tokens", None)
            forbrukning.append({
                "typ": "anrop",
                "datum": idag,
                "tidsstämpel": datetime.now().isoformat(timespec="seconds"),
                "ticker": ticker,
                "orsak": orsak,
                "modell": modell,
                "input_tokens": in_tok,
                "output_tokens": out_tok,
                "cache_creation_input_tokens": cache_skapad,
                "cache_read_input_tokens": cache_last,
                "körningsläge": körningsläge,
            })
        except anthropic.APIStatusError as e:
            print(f"    Claude-fel för {ticker}: {e.status_code} {e.message}")
        except Exception as e:
            print(f"    Claude-fel för {ticker}: {e}")
    return results, forbrukning


# ----------------------------------------------------------------------
# Claude-tokenförbrukning — en post per API-anrop, gist-synkad
# ----------------------------------------------------------------------
FORBRUKNING_FILE = "claude_forbrukning.json"
FORBRUKNING_KOMPRIMERA_TROSKEL = 5000   # rader — komprimera först när filen växer förbi detta
FORBRUKNING_KOMPRIMERA_ALDRE_AN_DAGAR = 90


def logga_forbrukning(poster):
    """Appenda dagens Claude-tokenposter till FORBRUKNING_FILE.

    Varje anropspost (se claude_analysis) har typ "anrop". Växer filen förbi
    FORBRUKNING_KOMPRIMERA_TROSKEL rader komprimeras anropsposter äldre än
    FORBRUKNING_KOMPRIMERA_ALDRE_AN_DAGAR dagar till veckosummor per modell
    (typ "veckosummering") så gisten inte sväller obegränsat. Ingenting
    krävs — no-op om poster är tom.
    """
    if not poster:
        return
    logg = []
    if os.path.exists(FORBRUKNING_FILE):
        try:
            with open(FORBRUKNING_FILE) as f:
                logg = json.load(f)
        except (json.JSONDecodeError, OSError):
            logg = []
    logg.extend(poster)
    if len(logg) > FORBRUKNING_KOMPRIMERA_TROSKEL:
        logg = _komprimera_forbrukning(logg)
    with open(FORBRUKNING_FILE, "w") as f:
        json.dump(logg, f, ensure_ascii=False, indent=2)


def _är_force_post(p):
    """En post räknas som "felsökning/force" om just den tickerns omanalys
    saknade en egen genuin trigger (orsak == "force-claude") — INTE samma sak
    som run-nivåns 'körningsläge'-fält, som taggar hela anropet som "force"
    så fort --force-claude användes, även för tickers med en egen riktig
    orsak (t.ex. "ny på listan"). De räknas här som normal drift, eftersom
    de skulle analyserats även utan flaggan."""
    if p.get("typ") == "veckosummering":
        return bool(p.get("force"))
    return p.get("orsak") == "force-claude"


def _komprimera_forbrukning(logg):
    """Slå ihop anropsposter äldre än FORBRUKNING_KOMPRIMERA_ALDRE_AN_DAGAR
    dagar till veckosummor (datum, antal_anrop, summa tokens) per modell och
    drift/force-status (se _är_force_post) så att --utvardera:s uppdelning
    fortsätter fungera även efter komprimering. Redan komprimerade
    veckosummeringar och färska anropsposter lämnas orörda.
    """
    from datetime import date, timedelta
    gräns = (date.today() - timedelta(days=FORBRUKNING_KOMPRIMERA_ALDRE_AN_DAGAR)).isoformat()
    behåll, komprimera = [], []
    for p in logg:
        if p.get("typ") == "anrop" and p.get("datum", "") < gräns:
            komprimera.append(p)
        else:
            behåll.append(p)
    if not komprimera:
        return logg

    per_vecka = {}
    for p in komprimera:
        d = date.fromisoformat(p["datum"])
        vecka_start = (d - timedelta(days=d.weekday())).isoformat()
        modell = p.get("modell") or "okänd"
        force = _är_force_post(p)
        agg = per_vecka.setdefault((vecka_start, modell, force),
                                   {"antal_anrop": 0, "input_tokens": 0, "output_tokens": 0})
        agg["antal_anrop"] += 1
        agg["input_tokens"] += p.get("input_tokens") or 0
        agg["output_tokens"] += p.get("output_tokens") or 0

    veckosummor = [
        {"typ": "veckosummering", "vecka_start": vecka, "modell": modell, "force": force, **agg}
        for (vecka, modell, force), agg in sorted(per_vecka.items())
    ]
    return veckosummor + behåll


def _veckointervall(datum_str):
    """(måndag, söndag) som ISO-strängar för veckan datum_str tillhör."""
    from datetime import date, timedelta
    d = date.fromisoformat(datum_str)
    måndag = d - timedelta(days=d.weekday())
    söndag = måndag + timedelta(days=6)
    return måndag.isoformat(), söndag.isoformat()


def skriv_forbrukningssammanfattning(claude_texts, forbrukning_denna_körning, idag):
    """Terminalsammanfattning efter Claude-steget: anrop/modell, återanvända,
    tokens denna körning + ackumulerat denna kalendervecka (mån–sön)."""
    antal_anrop = len(forbrukning_denna_körning)
    återanvända = len(claude_texts) - antal_anrop
    per_modell = {}
    for p in forbrukning_denna_körning:
        kort = MODELL_KORTNAMN.get(p.get("modell"), p.get("modell") or "okänd")
        per_modell[kort] = per_modell.get(kort, 0) + 1
    modell_text = ", ".join(f"{n} {namn}" for namn, n in per_modell.items())
    in_sum = sum(p.get("input_tokens") or 0 for p in forbrukning_denna_körning)
    out_sum = sum(p.get("output_tokens") or 0 for p in forbrukning_denna_körning)

    print(f"\nClaude-förbrukning denna körning:")
    print(f"  {antal_anrop} anrop ({modell_text}), {återanvända} återanvända"
          if modell_text else f"  {antal_anrop} anrop, {återanvända} återanvända")
    if antal_anrop:
        print(f"  Tokens denna körning: {in_sum} in / {out_sum} out")

    if not os.path.exists(FORBRUKNING_FILE):
        return
    try:
        with open(FORBRUKNING_FILE) as f:
            logg = json.load(f)
    except (json.JSONDecodeError, OSError):
        return
    måndag, söndag = _veckointervall(idag)
    veckoposter = [p for p in logg if p.get("typ") == "anrop" and måndag <= p.get("datum", "") <= söndag]
    if veckoposter:
        v_in = sum(p.get("input_tokens") or 0 for p in veckoposter)
        v_out = sum(p.get("output_tokens") or 0 for p in veckoposter)
        print(f"  Denna vecka ({måndag}–{söndag}): {len(veckoposter)} anrop, {v_in} in / {v_out} out tokens")


def _bygg_forbrukning_sektion(poster, med_fordelning=True):
    """Aggregerar en delmängd av loggen (drift ELLER force) till per_vecka
    (kombinerar råposter + redan komprimerade veckosummeringar) och, om
    begärt, per_orsak/per_modell (bara råposter — komprimerade veckor
    saknar den detaljen)."""
    per_vecka = {}
    for p in poster:
        if p.get("typ") == "veckosummering":
            key = p["vecka_start"]
            agg = per_vecka.setdefault(key, {"antal_anrop": 0, "tokens": 0})
            agg["antal_anrop"] += p.get("antal_anrop", 0)
            agg["tokens"] += (p.get("input_tokens") or 0) + (p.get("output_tokens") or 0)
        elif p.get("typ") == "anrop":
            måndag, _ = _veckointervall(p["datum"])
            agg = per_vecka.setdefault(måndag, {"antal_anrop": 0, "tokens": 0})
            agg["antal_anrop"] += 1
            agg["tokens"] += (p.get("input_tokens") or 0) + (p.get("output_tokens") or 0)

    total_antal = sum(v["antal_anrop"] for v in per_vecka.values())
    total_tokens = sum(v["tokens"] for v in per_vecka.values())
    resultat = {
        "per_vecka": per_vecka, "total_antal": total_antal, "total_tokens": total_tokens,
        "snitt_per_anrop": round(total_tokens / total_antal) if total_antal else None,
    }
    if med_fordelning:
        rå_poster = [p for p in poster if p.get("typ") == "anrop"]

        def _gruppera(nyckelfunk):
            grupper = {}
            for p in rå_poster:
                nyckel = nyckelfunk(p) or "okänd"
                agg = grupper.setdefault(nyckel, {"antal": 0, "tokens": 0})
                agg["antal"] += 1
                agg["tokens"] += (p.get("input_tokens") or 0) + (p.get("output_tokens") or 0)
            return grupper

        resultat["per_orsak"] = _gruppera(lambda p: p.get("orsak"))
        resultat["per_modell"] = _gruppera(lambda p: MODELL_KORTNAMN.get(p.get("modell"), p.get("modell")))
    return resultat


def claude_forbrukning_rapport():
    """Aggregerar FORBRUKNING_FILE för --utvardera, uppdelat i:
    - "drift": poster med en genuin trigger (orsak != "force-claude") —
      normal drift, huvudsiffran i rapporten.
    - "force": poster som bara tillkom pga. --force-claude-flaggan (ingen
      egen trigger) — felsökning/testkörningar, redovisas kompakt så de
      inte förorenar driftssiffrorna.
    Se _är_force_post för den exakta gränsdragningen (den skiljer sig
    medvetet från run-nivåns 'körningsläge'-fält).
    Returnerar None om filen saknas/är tom, annars {"drift": {...}, "force": {...}}.
    """
    if not os.path.exists(FORBRUKNING_FILE):
        return None
    try:
        with open(FORBRUKNING_FILE) as f:
            logg = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    if not logg:
        return None

    drift_poster = [p for p in logg if not _är_force_post(p)]
    force_poster = [p for p in logg if _är_force_post(p)]
    return {
        "drift": _bygg_forbrukning_sektion(drift_poster, med_fordelning=True),
        "force": _bygg_forbrukning_sektion(force_poster, med_fordelning=False),
    }


# ----------------------------------------------------------------------
# Steg 5: Ändringshistorik mellan körningar
# ----------------------------------------------------------------------
HISTORY_FILE = "portfolj_historik.json"
WEIGHT_CHANGE_THRESHOLD = 1.0  # procentenheter — mindre viktändringar loggas inte

# ----------------------------------------------------------------------
# Beständig lagring i GitHub Gist (för Render, vars disk är tillfällig).
# Aktiveras när GIST_ID och GITHUB_TOKEN är satta — annars no-op och allt
# fungerar lokalt som vanligt.
# ----------------------------------------------------------------------
GIST_ID = os.environ.get("GIST_ID")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
FACIT_FILE = "screener_facit.json"
PAPPER_FILE = "pappersportfolj.json"
# "Bästa köp" = konsensusaktier med poäng ≥ denna tröskel. 0 = alla
# konsensusaktier (nuvarande UI-beteende). Konstant för reproducerbara
# episoder (§3b Del A) — höj bara medvetet.
BASTA_KOP_MIN_POANG = 0
GIST_FILES = ("portfolj_historik.json", "senaste_analys.json",
              "bakgrund_topp50.json", "bakgrund_cache.json", FACIT_FILE, PAPPER_FILE,
              FORBRUKNING_FILE)


def _gist_headers():
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def gist_pull():
    """Hämta historik + senaste analys från gisten till lokala filer."""
    if not (GIST_ID and GITHUB_TOKEN):
        return
    try:
        r = requests.get(f"https://api.github.com/gists/{GIST_ID}",
                         headers=_gist_headers(), timeout=30)
        if r.status_code != 200:
            print(f"  OBS: kunde inte läsa gisten ({r.status_code}) — kör vidare utan.")
            return
        for name, f in (r.json().get("files") or {}).items():
            if name not in GIST_FILES or not f:
                continue
            content = f.get("content", "")
            if f.get("truncated"):
                content = requests.get(f["raw_url"], timeout=30).text
            if content.strip():
                with open(name, "w") as fh:
                    fh.write(content)
        print("  Historik hämtad från GitHub Gist.")
    except requests.RequestException as e:
        print(f"  OBS: gist-hämtning misslyckades ({e}) — kör vidare utan.")


def gist_push():
    """Spara historik + senaste analys till gisten."""
    if not (GIST_ID and GITHUB_TOKEN):
        return
    payload = {"files": {}}
    for name in GIST_FILES:
        if os.path.exists(name):
            with open(name) as fh:
                payload["files"][name] = {"content": fh.read()}
    if not payload["files"]:
        return
    try:
        r = requests.patch(f"https://api.github.com/gists/{GIST_ID}",
                           headers=_gist_headers(), json=payload, timeout=30)
        if r.status_code == 200:
            print("  Historik sparad till GitHub Gist.")
        else:
            print(f"  OBS: kunde inte spara till gisten ({r.status_code}).")
    except requests.RequestException as e:
        print(f"  OBS: gist-sparning misslyckades ({e}).")


def update_history(portfolios, consensus_tickers, near_tickers=None, exit_status=None):
    """Jämför med förra körningen, logga ändringar och spara ny ögonblicksbild.

    exit_status: {ticker: bool} — dagens exitvillkor (§B trendbrott) för
    tickers med giltig analys. Övergångar loggas EXIT (TRENDBROTT) / ÅTER
    FRÅN EXIT, med samma idempotensmönster som konsensus IN/UT (bara nya
    övergångar loggas, exit_datum bevaras oförändrat medan villkoret består).

    Returnerar (ändringslogg (nyaste först), exit_info {ticker: exit_datum}).
    """
    near_tickers = near_tickers or []
    exit_status = exit_status or {}
    from datetime import date
    today = date.today().isoformat()

    state = {"senaste": None, "logg": []}
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE) as f:
                state = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"  OBS: kunde inte läsa {HISTORY_FILE} ({e}) — börjar om historiken.")

    prev = state.get("senaste")
    entries = []

    def log(typ, profil, ticker, detalj):
        entries.append({"datum": today, "typ": typ, "profil": profil,
                        "ticker": ticker, "detalj": detalj})

    old_exit = (prev or {}).get("exit") or {}   # {ticker: exit_datum}
    today_exit_tickers = {tk for tk, flagga in exit_status.items() if flagga}
    current_tickers = set(consensus_tickers)
    for tk in sorted(today_exit_tickers - set(old_exit)):
        log("EXIT (TRENDBROTT)", "", tk, EXIT_VILLKOR_TEXT)
    for tk in sorted(set(old_exit) - today_exit_tickers):
        if tk in current_tickers:   # tyst borttagning om aktien lämnat konsensus (redan loggat)
            log("ÅTER FRÅN EXIT", "", tk, "Trend återhämtad — dödskorset inte längre aktivt")
    exit_info = {tk: old_exit.get(tk, today) for tk in today_exit_tickers}

    if prev:
        for profile, positions in portfolios.items():
            old = prev.get("portfolios", {}).get(profile)
            if old is None:
                log("NY PROFIL", profile, "", f"{len(positions)} innehav")
                continue
            for tk, w in positions.items():
                if tk not in old:
                    log("NYTT INNEHAV", profile, tk, f"vikt {w} %")
                elif abs(w - old[tk]) >= WEIGHT_CHANGE_THRESHOLD:
                    log("VIKTÄNDRING", profile, tk, f"{old[tk]} % → {w} %")
            for tk, w in old.items():
                if tk not in positions:
                    log("SÅLT INNEHAV", profile, tk, f"hade vikt {w} %")

        old_cons = set(prev.get("consensus", []))
        new_cons = set(consensus_tickers)
        old_near = set(prev.get("near_consensus") or [])
        new_near = set(near_tickers)

        n = len(portfolios)
        in_krav, kvar_krav = konsensus_trosklar(n)
        in_pct = round(KONSENSUS_ANDEL_IN * 100)
        kvar_pct = round(KONSENSUS_ANDEL_KVAR * 100)

        def agare(tk):
            return sum(1 for p in portfolios.values() if tk in p)

        for tk in sorted(new_cons - old_cons):
            detalj = f"klarar {in_pct} %-innivån ({agare(tk)} av {n} portföljer)"
            if tk in old_near:
                detalj += " — upp från nära konsensus"
            log("IN I KONSENSUS", "", tk, detalj)
        for tk in sorted(old_cons - new_cons):
            detalj = f"under {kvar_pct} %-kvarnivån ({agare(tk)} av {n} portföljer)"
            if tk in new_near:
                detalj += " — ner till nära konsensus"
            log("UT UR KONSENSUS", "", tk, detalj)

        # Nära konsensus-övergångar loggas bara om förra ögonblicksbilden
        # skrevs under SAMMA regelversion — annars skulle en omdefinition av
        # nivån (t.ex. bytet från "2 av 5" till kvarnivån) spamma loggen och
        # "Lämnat listorna" med falska utträden
        if prev.get("regel") == KONSENSUS_REGEL:
            for tk in sorted(new_near - old_near - old_cons):
                log("IN I NÄRA KONSENSUS", "", tk,
                    f"klarar kvarnivåns antal ({agare(tk)} av {n} portföljer)")
            for tk in sorted(old_near - new_near - new_cons):
                log("UT UR NÄRA KONSENSUS", "", tk,
                    f"under kvarnivån ({agare(tk)} av {n} portföljer)")

        if entries:
            print(f"  {len(entries)} ändringar sedan {prev.get('datum', 'förra körningen')}.")
        else:
            print(f"  Inga ändringar sedan {prev.get('datum', 'förra körningen')}.")
    else:
        print("  Första körningen — skapar utgångsläge för historiken.")

    state["senaste"] = {"datum": today, "portfolios": portfolios,
                        "consensus": sorted(consensus_tickers),
                        "near_consensus": sorted(near_tickers),
                        "regel": KONSENSUS_REGEL,
                        "exit": exit_info}
    state["logg"] = entries + state.get("logg", [])
    with open(HISTORY_FILE, "w") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

    return state["logg"], exit_info


# ----------------------------------------------------------------------
# Steg 6: Excel-rapport
# ----------------------------------------------------------------------
def write_excel(portfolios, consensus, analyses, claude_texts, history_log,
                ranking=None, near_consensus=None, holding_info=None, divergence=None,
                divergence_near=None, bubblar_niva=None, regim=None, exit_lista=None):
    from datetime import date, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    holding_info = holding_info or {}
    nyhets_cutoff = (date.today() - timedelta(days=7)).isoformat()
    rank_by_ticker = {r["ticker"]: r for r in (ranking or [])}

    def is_new(ticker, typ):
        return any(e["ticker"] == ticker and e["typ"] == typ and e["datum"] >= nyhets_cutoff
                   for e in history_log or [])

    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1F4E78")

    def style_header(ws):
        for c in ws[1]:
            c.font, c.fill = hfont, hfill

    # Rangordning — sammanvägd poäng, bästa köp först
    if ranking or exit_lista:
        green = PatternFill("solid", start_color="C6EFCE")
        red = PatternFill("solid", start_color="FFC7CE")
        exit_fill = PatternFill("solid", start_color="FFD966")
        ws = wb.create_sheet("Rangordning", 0)
        if regim:
            badge = (f"Marknadsregim: {regim.get('regim')} "
                     f"({regim.get('regim_kalla') or 'index'} {regim.get('spy_pris')} "
                     f"vs MA200 {regim.get('spy_ma200')})")
            if regim.get("notis"):
                badge += f"  — {regim['notis']}"
            ws.append([badge])
            varning_cell = ws.cell(row=1, column=1)
            varning_cell.font = Font(bold=True, size=12,
                                     color="9C0006" if (regim.get("notis") or "").startswith("⚠") else "000000")
            ws.append([])
        ws.append(["Rang", "Instrument", "Poäng (0–100)", "Poäng (v1)", "Stigande trend",
                   "Trend (25)", "Momentum (20)", "Analytiker (20)", "Konsensus (25)",
                   "Värdering (10)", "Relativ styrka (pe)", "Föreslagen vikt (%)",
                   "Claudes rekommendation"])
        for c in ws[ws.max_row]:
            c.font, c.fill = hfont, hfill
        for i, r in enumerate(ranking or [], start=1):
            d = r["delpoäng"]
            rs = r.get("relativ_styrka") or {}
            c_rek = claude_texts.get(r["ticker"], {})
            ws.append([i, r["ticker"], r["poäng"], r.get("poäng_v1"),
                       "JA" if r["trend_ok"] else "NEJ",
                       d.get("Trend"), d.get("Momentum"), d.get("Analytiker"), d.get("Konsensus"),
                       d.get("Värdering"), rs.get("rs_pe"), r.get("foreslagen_vikt_%"),
                       c_rek.get("rekommendation_visning") or c_rek.get("rekommendation", "")])
            ws.cell(row=ws.max_row, column=5).fill = green if r["trend_ok"] else red
            if r.get("vardering_neutral"):
                ws.cell(row=ws.max_row, column=10).comment = Comment(
                    "Data saknas → neutral (5/10)", "System")

        if exit_lista:
            ws.append([])
            ws.append([f"EXIT (TRENDBROTT) — {len(exit_lista)} st, uteslutna ur Bästa köp "
                       "(kvar i konsensus, se Konsensus & Analys)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append(["Instrument", "Exit-datum", "Villkor", "Poäng vid exit"])
            for c in ws[ws.max_row]:
                if c.value:
                    c.font, c.fill = hfont, hfill
            for r in exit_lista:
                ws.append([r["ticker"], r["exit_datum"], r["exit_villkor"], r["poäng"]])
                for col in range(1, 5):
                    ws.cell(row=ws.max_row, column=col).fill = exit_fill

    # Flik per profil
    for name, positions in portfolios.items():
        ws = wb.create_sheet(name[:31])
        ws.append(["Instrument", "Vikt (%)"])
        style_header(ws)
        for ticker, w in sorted(positions.items(), key=lambda x: -x[1]):
            ws.append([ticker, w])

    # Konsensus + analys
    green = PatternFill("solid", start_color="C6EFCE")
    red = PatternFill("solid", start_color="FFC7CE")
    # Divergens — signalgruppens unika övertygelser vs flocken
    if divergence:
        ws = wb.create_sheet("Divergens", 1)
        ws.append(["Instrument", "I signalgrupp", "Signalandel (%)", "I bakgrund (antal)",
                   "Bakgrundsandel (%)", "Divergens (pp)", "Bakgrundens snittvikt (%)",
                   "Claudes rekommendation", "Sektor", "Kluster-ID", "Klusterfaktor"])
        style_header(ws)
        for tk, dv in sorted(divergence.items(), key=lambda x: -x[1]["divergens_pp"]):
            kluster = rank_by_ticker.get(tk, {}).get("kluster") or {}
            ws.append([tk, f"{dv['signal_antal']}/{len(portfolios)}", dv["signal_andel_pct"],
                       dv["bakgrund_antal"], dv["bakgrund_andel_pct"], dv["divergens_pp"],
                       dv["bakgrund_snittvikt"],
                       claude_texts.get(tk, {}).get("rekommendation", ""),
                       analyses.get(tk, {}).get("sector"),
                       kluster.get("kluster_id"), kluster.get("klusterfaktor")])

        # Bubblare: bubblarnivån med hög divergens
        bubbel_kalla = bubblar_niva if bubblar_niva is not None else (near_consensus or {})
        bubblare = sorted(((tk, dv) for tk, dv in (divergence_near or {}).items()
                           if dv["divergens_pp"] >= 30),
                          key=lambda x: -x[1]["divergens_pp"])
        if bubblare:
            ws.append([])
            ws.append(["BUBBLARE — bubblarnivån med hög divergens (nära att kvala in)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append(["Instrument", "Ägs av (signalgrupp)", "Total vikt (%)",
                       "I bakgrund (antal)", "Bakgrundsandel (%)", "Divergens (pp)"])
            for c in ws[ws.max_row]:
                if c.value:
                    c.font, c.fill = hfont, hfill
            for tk, dv in bubblare:
                info = bubbel_kalla.get(tk, {})
                ws.append([tk, ", ".join(info.get("holders", [])),
                           info.get("total_weight"),
                           dv["bakgrund_antal"], dv["bakgrund_andel_pct"], dv["divergens_pp"]])

    ws = wb.create_sheet("Konsensus & Analys", 2)
    ws.append(["Instrument", "Ny", "Stigande trend", "Antal portföljer", "Snittvikt (%)", "Pris", "RSI14",
               "Över MA200", "Rekommendation", "Riktkurs", "Uppsida (%)", "Antal analytiker",
               "Ägd längst (dagar)", "Investerarnas snittvinst (%)",
               "Viktad konsensus", "Senaste köp (dagar)", "EPS-rev 90d (%)", "Nettoflöde 30d (pe)",
               "P/E (forward)", "PEG", "Riktkurs spridning", "Nästa rapport"])
    style_header(ws)
    varning_fill = PatternFill("solid", start_color="FFEB9C")
    idag = date.today()
    consensus_order = sorted(consensus.items(), key=lambda x: (-x[1]["count"], -x[1]["avg_weight"]))
    for ticker, info in consensus_order:
        a = analyses.get(ticker, {})
        h = holding_info.get(ticker, {})
        trend = a.get("stigande_trend")
        spridning = a.get("riktkurs_spridningskvot")
        hog, lag = a.get("riktkurs_hog"), a.get("riktkurs_lag")
        spridning_txt = f"{lag:g}–{hog:g}" if (hog and lag) else None
        rapport = a.get("nasta_rapport")
        rapport_snart = False
        if rapport:
            try:
                rapport_snart = (date.fromisoformat(rapport) - idag).days <= 7
            except ValueError:
                pass
        ws.append([
            ticker, "NY" if is_new(ticker, "IN I KONSENSUS") else "",
            "JA" if trend else ("NEJ" if trend is not None else "?"),
            info["count"], round(info["avg_weight"], 2),
            a.get("pris"), a.get("RSI14"), a.get("över_MA200"),
            a.get("rekommendation"), a.get("riktkurs"), a.get("uppsida_%"), a.get("antal_analytiker"),
            h.get("längst_dagar"), h.get("snitt_vinst_pct"),
            info.get("viktad_konsensus"), info.get("senaste_köp_dagar"),
            a.get("eps_rev_90d_pct"),
            rank_by_ticker.get(ticker, {}).get("nettoflode_30d_pe"),
            a.get("forward_pe"), a.get("peg_ratio"), spridning_txt,
            f"⚠ {rapport}" if rapport_snart else rapport,
        ])
        if trend is not None:
            ws.cell(row=ws.max_row, column=3).fill = green if trend else red
        if rapport_snart:
            ws.cell(row=ws.max_row, column=ws.max_column).fill = varning_fill

    # Nära konsensus — en portfölj från att kvala in
    if near_consensus:
        ws.append([])
        _in_krav, _kvar_krav = konsensus_trosklar(len(portfolios))
        ws.append([f"NÄRA KONSENSUS — {_kvar_krav} av {len(portfolios)} portföljer "
                   f"(klarar kvarnivån men inte innivån {_in_krav})"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Instrument", "Ny", "Antal portföljer", "Total vikt (%)", "Snittvikt (%)",
                   "Ägs av", "Ägd längst (dagar)", "Investerarnas snittvinst (%)"])
        for c in ws[ws.max_row]:
            if c.value:
                c.font, c.fill = hfont, hfill
        def _total(info):
            return info.get("total_weight") or round(info["avg_weight"] * info["count"], 2)
        for ticker, info in sorted(near_consensus.items(), key=lambda x: -_total(x[1])):
            h = holding_info.get(ticker, {})
            ws.append([ticker, "NY" if is_new(ticker, "IN I NÄRA KONSENSUS") else "",
                       info["count"], _total(info), round(info["avg_weight"], 2),
                       ", ".join(info.get("holders", [])),
                       h.get("längst_dagar"), h.get("snitt_vinst_pct")])

    # Lämnat listorna nyligen (senaste 30 dagarna)
    lamnat_cutoff = (date.today() - timedelta(days=30)).isoformat()
    lamnat = [e for e in history_log or []
              if e["typ"] in ("UT UR KONSENSUS", "UT UR NÄRA KONSENSUS")
              and e["datum"] >= lamnat_cutoff]
    if lamnat:
        ws.append([])
        ws.append(["LÄMNAT LISTORNA (senaste 30 dagarna)"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Datum", "Instrument", "Typ", "Detalj"])
        for c in ws[ws.max_row]:
            if c.value:
                c.font, c.fill = hfont, hfill
        for e in lamnat:
            ws.append([e["datum"], e["ticker"], e["typ"], e["detalj"]])

    # Teknisk analys (indikatorer + Claudes text)
    exit_by_ticker = {r["ticker"]: r for r in (exit_lista or [])}
    today_str = date.today().isoformat()
    ws = wb.create_sheet("Teknisk analys", 3)
    ws.append(["Instrument", "Stigande trend", "Över MA200", "MA200 stigande",
               "Pris", "MA50", "MA200", "Golden cross", "RSI14",
               "MACD > signal", "Bollinger (%)", "Avstånd 52v-högsta (%)",
               "Avkastning 1m (%)", "Avkastning 3m (%)", "Volymtrend (%)", "EXIT (trendbrott)",
               "Claudes rekommendation", "Claudes analys", "Analysstatus", "Modell",
               "EPS-rev 90d (%)"])
    style_header(ws)
    for ticker, _ in consensus_order:
        a = analyses.get(ticker, {})
        c = claude_texts.get(ticker, {})
        trend = a.get("stigande_trend")
        exit_r = exit_by_ticker.get(ticker)
        gen = c.get("genererad")
        if gen == today_str:
            status = f"Analys från {gen} (ny analys, orsak: {c.get('analys_orsak') or '?'})"
        elif gen:
            status = f"Analys från {gen} (återanvänd, orsak: inga väsentliga ändringar)"
        else:
            status = ""
        ws.append([
            ticker, "JA" if trend else ("NEJ" if trend is not None else "?"),
            a.get("över_MA200"), a.get("MA200_stigande"),
            a.get("pris"), a.get("MA50"), a.get("MA200"), a.get("golden_cross"),
            a.get("RSI14"), a.get("MACD_över_signal"), a.get("bollinger_position_%"),
            a.get("avstånd_52v_högsta_%"), a.get("avkastning_1m_%"), a.get("avkastning_3m_%"),
            a.get("volymtrend_20d_vs_3m_%"),
            f"sedan {exit_r['exit_datum']}" if exit_r else "",
            c.get("rekommendation_visning") or c.get("rekommendation"), c.get("analys"), status,
            c.get("modell"),
            a.get("eps_rev_90d_pct"),
        ])
        if trend is not None:
            ws.cell(row=ws.max_row, column=2).fill = green if trend else red
        if exit_r:
            ws.cell(row=ws.max_row, column=16).fill = PatternFill("solid", start_color="FFD966")
    ws.column_dimensions["R"].width = 100
    for row in ws.iter_rows(min_row=2, min_col=18, max_col=18):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Historik (ändringslogg, nyaste först)
    ws = wb.create_sheet("Historik", 4)
    ws.append(["Datum", "Typ", "Profil", "Instrument", "Detalj"])
    style_header(ws)
    if history_log:
        for e in history_log:
            ws.append([e["datum"], e["typ"], e["profil"], e["ticker"], e["detalj"]])
    else:
        ws.append(["", "Första körningen — ändringar visas från nästa körning.", "", "", ""])
    for col, width in (("A", 12), ("B", 18), ("C", 18), ("D", 12), ("E", 40)):
        ws.column_dimensions[col].width = width

    del wb["Sheet"]
    wb.save(OUTPUT_FILE)
    print(f"\nKlart! Rapport sparad som: {OUTPUT_FILE}")


def excel_from_result(result):
    """Återskapa portfolj_analys.xlsx från ett sparat resultat (RESULTS_FILE).

    Används av webbappen så att nedladdningsknappen alltid fungerar — även
    när Excel-filen saknas (t.ex. efter att Render vaknat ur sömn utan att
    ha kört en ny analys). All data finns redan i resultatet.
    """
    write_excel(
        result.get("portfolios", {}),
        result.get("consensus", {}),
        result.get("analyses", {}),
        result.get("claude", {}),
        result.get("historik", []),
        result.get("ranking"),
        result.get("nara_konsensus"),
        result.get("innehav"),
        result.get("divergens"),
        result.get("divergens_nara"),
        result.get("bubblar_niva"),
        result.get("regim"),
        result.get("exit_lista"),
    )


# ----------------------------------------------------------------------
# Huvudflöde
# ----------------------------------------------------------------------
RESULTS_FILE = "senaste_analys.json"


def run_analysis(with_claude=True, force_claude=False, refresh_background=False):
    """Kör hela pipelinen. Returnerar resultatet som dict och sparar det
    till RESULTS_FILE + portfolj_analys.xlsx. Kastar RuntimeError vid fel
    (så att webbappen kan visa felet i stället för att dö).

    Claude-analysen körs max en gång per dag (kostar API-credits) — har den
    redan körts idag återanvänds texterna. Nya konsensusaktier analyseras
    dock alltid. force_claude=True kringgår dagsspärren.

    refresh_background=True (--divergens) hämtar om bakgrundsgruppens
    portföljer; annars används befintlig cache utan extra API-anrop.
    """
    from datetime import datetime, date

    # Körningsläge för tokenförbrukningsloggen (se logga_forbrukning)
    körningsläge = "force" if force_claude else ("divergens" if refresh_background else "standard")

    if not API_KEY or not USER_KEY:
        raise RuntimeError("ETORO_API_KEY och ETORO_USER_KEY saknas — lägg dem i .env-filen.")

    # Hämta beständig historik (Render har tillfällig disk)
    gist_pull()

    # §A Marknadsregim (SPY vs MA200) — oberoende av eToro-data, 1 yfinance-anrop
    regim = compute_market_regime()
    print(f"\nMarknadsregim: {regim['regim']} (SPY {regim['spy_pris']} vs MA200 {regim['spy_ma200']})"
          + (f"  — {regim['notis']}" if regim.get("notis") else ""))

    print("\nTestar API-anslutning...")
    test = api_get("/market-data/search", {"query": "Apple"})
    if test is None:
        raise RuntimeError("eToro-anslutningen misslyckades. Kontrollera nycklarna i .env.")
    print("Anslutning OK!")

    portfolios = {}
    port_meta = {}
    for username in PROFILES:
        p, meta = get_portfolio(username)
        if p:
            portfolios[username] = p
            port_meta[username] = meta or {}

    if not portfolios:
        raise RuntimeError("Inga portföljer kunde hämtas via eToro-API:et.")

    # Konsensus (procentuella trösklar med hysteres) + nära konsensus + bubblarnivå
    previous_consensus = load_previous_consensus()
    consensus, near_consensus, bubblar_niva = compute_consensus(
        portfolios, port_meta, previous_consensus)

    in_krav, kvar_krav = konsensus_trosklar(len(portfolios))
    print(f"\nKonsensus-aktier (in: {in_krav} av {len(portfolios)}, "
          f"kvar: {kvar_krav} av {len(portfolios)}): {sorted(consensus)}")
    print(f"Nära konsensus ({kvar_krav} ägare, under innivån): {len(near_consensus)} st")
    print(f"Bubblarnivå ({kvar_krav - 1} ägare): {len(bubblar_niva)} st")

    # Bakgrundsgrupp + divergens (vad äger signalgruppen som flocken INTE äger?)
    background = load_background_portfolios(refresh=refresh_background)
    divergence = compute_divergence(consensus, portfolios, background)
    # Bubblare: bubblarnivå-aktier med hög divergens — nära att kvala in,
    # och flocken äger dem inte
    divergence_near = compute_divergence(bubblar_niva, portfolios, background)
    if divergence:
        rankad = sorted(divergence.items(), key=lambda x: -x[1]["divergens_pp"])
        print("Divergens (signalgrupp − bakgrundsgrupp):")
        for tk, dv in rankad:
            print(f"  {tk:<6} signal {dv['signal_andel_pct']}% | "
                  f"bakgrund {dv['bakgrund_andel_pct']}% | divergens {dv['divergens_pp']:+.1f} pp")

    # Förra körningens resultat (för Claude-dagsspärren och som reserv
    # för analytikerdata när Yahoo blockerar)
    prev = {}
    if os.path.exists(RESULTS_FILE):
        try:
            with open(RESULTS_FILE) as f:
                prev = json.load(f)
        except (json.JSONDecodeError, OSError):
            prev = {}

    # Regimkedjan (SPY/^GSPC/VOO/IVV) kan blockeras helt (samma Yahoo-symptom
    # som analysfälten ovan) — återanvänd förra kända regimen hellre än OKÄND.
    # regim_datum (datum för senaste LYCKADE beräkning, ej senaste körning)
    # följer med oförändrat — en åldersvarning eskaleras om den blir för gammal.
    if regim["regim"] == "OKÄND":
        prev_regim = prev.get("regim") or {}
        if prev_regim.get("regim") and prev_regim["regim"] != "OKÄND":
            regim_datum = prev_regim.get("regim_datum") or prev_regim.get("datum")
            alder = _handelsdagar_mellan(regim_datum, date.today().isoformat()) if regim_datum else None
            if alder is not None and alder > REGIM_ALDER_VARNING_HANDELSDAGAR:
                notis = f"⚠ regim baserad på {alder} dagar gammal data"
            else:
                notis = f"återanvänd — {regim.get('notis')}"
            print(f"    Regim återanvänd från {regim_datum or 'okänt datum'} "
                  f"({alder if alder is not None else '?'} handelsdagar sedan): "
                  f"{prev_regim['regim']} — {notis}")
            regim = {**prev_regim, "regim_datum": regim_datum, "notis": notis}

    # Teknisk analys + analytikerdata
    import time
    analyses = {}
    for ticker in consensus:
        print(f"Analyserar {ticker}...")
        analyses[ticker] = analyze_ticker(ticker)
        if "error" in analyses[ticker]:
            print(f"    FEL för {ticker}: {analyses[ticker]['error']}")
        elif analyses[ticker].get("datakälla") != "Yahoo":
            print(f"    (Yahoo blockerade — data från {analyses[ticker]['datakälla']})")
        time.sleep(1.5)   # snäll paus så Yahoo inte strypmarkerar oss

    # Fallerar båda datakällorna (t.ex. Alpha Vantages dagsgräns på 25 anrop)?
    # Återanvänd då hela förra körningens analys i stället för en tom rad.
    prev_analyses = prev.get("analyses") or {}
    for ticker, a in analyses.items():
        if "error" not in a:
            continue
        pa = prev_analyses.get(ticker)
        if pa and "error" not in pa:
            cached = dict(pa)
            cached["datakälla"] = "cache"
            cached["cache_datum"] = (prev.get("tidpunkt") or "")[:16].replace("T", " kl. ")
            analyses[ticker] = cached
            print(f"    {ticker}: datakällorna svarade inte — återanvänder analysen från {cached['cache_datum']}")
    for ticker, a in analyses.items():
        if "error" in a or a.get("riktkurs"):
            continue
        pa = prev_analyses.get(ticker) or {}
        if pa.get("riktkurs"):
            a["rekommendation"] = pa.get("rekommendation", "n/a")
            a["riktkurs"] = pa["riktkurs"]
            a["antal_analytiker"] = pa.get("antal_analytiker")
            if a.get("pris"):
                a["uppsida_%"] = round((pa["riktkurs"] / a["pris"] - 1) * 100, 1)
            print(f"    {ticker}: analytikerdata återanvänd från {prev.get('tidpunkt', 'förra körningen')[:10]}")

    # Yahoo kan blockera .info/.eps_trend/.calendar separat från historik/pris
    # (felsökt 2026-07-17 — Render-specifikt, fälten fylls lokalt). Fältvis
    # återanvändning från förra körningen, oberoende av riktkurs-fallbacken ovan.
    ATERANVANDBARA_INFO_FALT = [
        "eps_rev_90d_pct", "forward_pe", "peg_ratio", "riktkurs_hog", "riktkurs_lag",
        "riktkurs_spridningskvot", "nasta_rapport", "sector", "industry",
    ]
    for ticker, a in analyses.items():
        if "error" in a:
            continue
        pa = prev_analyses.get(ticker) or {}
        återanvänt = [f for f in ATERANVANDBARA_INFO_FALT
                     if a.get(f) is None and pa.get(f) is not None]
        for falt in återanvänt:
            a[falt] = pa[falt]
        if återanvänt:
            print(f"    {ticker}: {len(återanvänt)} värderings-/rapportfält återanvända från "
                  f"{prev.get('tidpunkt', 'förra körningen')[:10]} ({', '.join(återanvänt)})")

    # §B Exitregel (trendbrott): pris < MA200 OCH MA50 < MA200 — beräknas här
    # (kräver bara analyses) men appliceras i update_history/build_ranking nedan
    exit_status = compute_exit_status(analyses, [tk for tk in consensus if "error" not in analyses.get(tk, {})])

    # Innehavstid + investerarnas upparbetade vinst (från positionernas
    # openTimestamp/netProfit) för konsensus- och nära konsensus-aktier
    holding_info = {}
    today_d = date.today()
    for ticker in list(consensus) + list(near_consensus) + list(bubblar_niva):
        per_profil = {}
        for prof, meta in port_meta.items():
            m = meta.get(ticker) or {}
            if m.get("öppnad"):
                dagar = (today_d - date.fromisoformat(m["öppnad"])).days
                per_profil[prof] = {"dagar": dagar, "vinst_pct": m.get("vinst_pct")}
        if not per_profil:
            continue
        längst_prof, längst = max(per_profil.items(), key=lambda x: x[1]["dagar"])
        vinster = [v["vinst_pct"] for v in per_profil.values() if v["vinst_pct"] is not None]
        holding_info[ticker] = {
            "längst_dagar": längst["dagar"],
            "längst_profil": längst_prof,
            "snitt_dagar": round(sum(v["dagar"] for v in per_profil.values()) / len(per_profil)),
            "snitt_vinst_pct": round(sum(vinster) / len(vinster), 1) if vinster else None,
            "per_profil": per_profil,
        }

    # Ge Claude innehavskontexten (för "ta hem vinsten"-bedömningen)
    # och divergensen (unik övertygelse vs flockbeteende)
    for ticker, a in analyses.items():
        h = holding_info.get(ticker)
        if h and "error" not in a:
            a["investerarnas_innehavstid_dagar_längst"] = h["längst_dagar"]
            a["investerarnas_innehavstid_dagar_snitt"] = h["snitt_dagar"]
            a["investerarnas_upparbetade_vinst_pct_snitt"] = h["snitt_vinst_pct"]
        dv = divergence.get(ticker)
        if dv and "error" not in a:
            a["ägs_av_pct_av_bakgrundsgruppen"] = dv["bakgrund_andel_pct"]
            a["divergens_mot_bakgrundsgruppen_pp"] = dv["divergens_pp"]

    # Ändringshistorik mot förra körningen (görs innan rangordningen så
    # nettoflödet §5 kan använda dagens nyloggade ändringar också). Tar även
    # med exit_status så EXIT (TRENDBROTT)/ÅTER FRÅN EXIT loggas (§B). Flyttad
    # hit (före Claude-steget) så dagens poäng/exit-status finns tillgängliga
    # för Claude-triggerfiltret nedan.
    print("\nUppdaterar ändringshistorik...")
    history_log, exit_info = update_history(portfolios, list(consensus.keys()),
                                             list(near_consensus.keys()), exit_status)

    # Sammanvägd rangordning (exit_lista = Bästa köp-kandidater i EXIT, §B)
    ranking, exit_lista = build_ranking(analyses, consensus, history_log, exit_info)
    print("\nRangordning (bästa köp först):")
    for i, r in enumerate(ranking, start=1):
        print(f"  {i}. {r['ticker']}: {r['poäng']} p (trend {'OK' if r['trend_ok'] else 'EJ OK'})")
    if exit_lista:
        print(f"EXIT (trendbrott), uteslutna ur Bästa köp: "
              f"{', '.join(r['ticker'] for r in exit_lista)}")
    poang_by_ticker = {r["ticker"]: r["poäng"] for r in ranking + exit_lista}
    delpoang_by_ticker = {r["ticker"]: r["delpoäng"] for r in ranking + exit_lista}

    # Claude skriver en gedigen teknisk analys per konsensusaktie
    # (max en gång per dag — återanvänd dagens texter om de finns). Steg 4:
    # Claude-triggerfilter — inom dagsspärren/helgvilan analyseras bara
    # aktier med väsentliga förändringar sedan texten skrevs om.
    prev_claude = prev.get("claude") or {}
    prev_datum = prev.get("claude_datum")
    today = date.today().isoformat()

    def _bygg_jobb(tickers, force_alla):
        jobb = {}
        for tk in tickers:
            a = analyses.get(tk, {})
            if "error" in a:
                continue
            cons = consensus.get(tk, {})
            snapshot = _bygg_indikator_snapshot(
                a, poang_by_ticker.get(tk), cons.get("viktad_konsensus"), tk in exit_info,
                regim["regim"])
            prev_entry = prev_claude.get(tk)
            ny = not prev_entry or not prev_entry.get("analys")
            if force_alla:
                behövs, orsak = True, ("ny på listan" if ny else "force-claude")
            else:
                behövs, orsak = behover_ny_analys(tk, snapshot, prev_entry)
            if behövs:
                dv = divergence.get(tk, {})
                claude_input = _bygg_claude_input(
                    tk, a, poang_by_ticker.get(tk), delpoang_by_ticker.get(tk),
                    cons.get("viktad_konsensus"), dv.get("divergens_pp"),
                    tk in exit_info)
                jobb[tk] = {"data": claude_input, "orsak": orsak, "snapshot": snapshot,
                           "model": CLAUDE_MODELL_NY if ny else CLAUDE_MODELL_OMANALYS}
        return jobb

    claude_texts = {}
    claude_datum = None
    forbrukning_denna_körning = []
    helg = date.today().weekday() in (5, 6)   # lördag/söndag — marknaden stängd
    if with_claude:
        if helg and prev_claude and not force_claude:
            # Helgvila: kurserna är fredagens, så senaste analysen gäller ännu
            claude_texts = {t: c for t, c in prev_claude.items() if t in consensus}
            claude_datum = prev_datum
            print("\nHelg — marknaden stängd, Claude-analysen från senaste handelsdagen återanvänds.")
        elif prev_datum == today and not force_claude:
            claude_texts = {t: c for t, c in prev_claude.items() if t in consensus}
            claude_datum = prev_datum
            jobb = _bygg_jobb([t for t in analyses if t not in claude_texts], force_alla=False)
            if jobb:
                print(f"\nClaude-analys körd idag — analyserar {len(jobb)} nya aktier...")
                nya_texter, forbrukning_denna_körning = claude_analysis(jobb, körningsläge)
                claude_texts.update(nya_texter)
            else:
                print("\nClaude-analysen är redan körd idag — återanvänder den (max 1 gång/dag).")
        else:
            claude_texts = {t: c for t, c in prev_claude.items() if t in consensus}
            jobb = _bygg_jobb(list(analyses.keys()), force_alla=force_claude)
            if jobb:
                verb = "Force-claude: analyserar alla" if force_claude else \
                       f"Claude-analys: {len(jobb)} aktier med väsentliga förändringar"
                print(f"\n{verb} (av {len(analyses)})...")
                nya_texter, forbrukning_denna_körning = claude_analysis(jobb, körningsläge)
                claude_texts.update(nya_texter)
            else:
                print("\nInga väsentliga förändringar sedan senaste Claude-analysen — allt återanvänds.")
            claude_datum = today if claude_texts else None
    else:
        # Claude bortvald denna körning — visa senaste kända texter
        claude_texts = {t: c for t, c in prev_claude.items() if t in consensus}
        claude_datum = prev_datum

    # analys_alder_dagar (för visning: "Analys från ... (återanvänd, ...)")
    # sätts för samtliga texter, oavsett om de precis skrevs eller återanvänds
    for c in claude_texts.values():
        gen = c.get("genererad")
        c["analys_alder_dagar"] = (today_d - date.fromisoformat(gen)).days if gen else None

    # Tokenförbrukning: logga dagens anrop + sammanfatta denna körning + veckan
    if with_claude:
        logga_forbrukning(forbrukning_denna_körning)
        skriv_forbrukningssammanfattning(claude_texts, forbrukning_denna_körning, today)

    # §A: i RÖD regim nedgraderas Claudes KÖP-text på Bästa köp-aktier till
    # visning ("KÖP (vänta på marknaden)") — poängen/claude_rek i facit
    # förblir OFÖRÄNDRADE så mätserierna inte förorenas av regimfiltret.
    bästa_köp_tickers = {r["ticker"] for r in ranking}
    for tk, c in claude_texts.items():
        rek = c.get("rekommendation")
        if regim["regim"] == "RÖD" and rek == "KÖP" and tk in bästa_köp_tickers:
            c["rekommendation_visning"] = "KÖP (vänta på marknaden)"
        else:
            c["rekommendation_visning"] = rek

    result = {
        "tidpunkt": datetime.now().isoformat(timespec="minutes"),
        "profiler": list(portfolios.keys()),
        "portfolios": portfolios,
        "consensus": consensus,
        "nara_konsensus": near_consensus,
        "bubblar_niva": bubblar_niva,
        "konsensus_trosklar": {"in": in_krav, "kvar": kvar_krav,
                               "n": len(portfolios),
                               "in_pct": round(KONSENSUS_ANDEL_IN * 100),
                               "kvar_pct": round(KONSENSUS_ANDEL_KVAR * 100)},
        "analyses": analyses,
        "claude": claude_texts,
        "claude_datum": claude_datum,
        "ranking": ranking,
        "exit_lista": exit_lista,
        "regim": regim,
        "historik": history_log,
        "innehav": holding_info,
        "divergens": divergence,
        "divergens_nara": divergence_near,
        "bakgrund_antal": len(background) if background else 0,
        "bransch": {tk: _industry_cache.get(tk)
                    for tk in list(consensus) + list(near_consensus) + list(bubblar_niva)},
    }
    with open(RESULTS_FILE, "w") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    write_excel(portfolios, consensus, analyses, claude_texts, history_log, ranking,
                near_consensus, holding_info, divergence, divergence_near, bubblar_niva,
                regim, exit_lista)

    # Logga facit för --utvardera (poäng vs framtida avkastning) — ranking
    # innehåller redan inte EXIT-aktier, så facit/episoder påverkas naturligt
    logga_facit(today, ranking, divergence, analyses, consensus, claude_texts)

    # Logga dagens pappersportfölj-vikter (§3b Del B — ombalansering; regim
    # styr §A:s "inga nya köp i RÖD", ranking utesluter redan EXIT-aktier)
    logga_pappersportfolj(today, ranking, claude_texts, regim)

    # Spara beständig historik till gisten
    gist_push()
    return result


def logga_facit(datum, ranking, divergence, analyses, consensus, claude_texts=None):
    """Appenda dagens poäng/divergens per konsensusaktie till FACIT_FILE.

    En rad per aktie och dag (idempotent — samma dag skrivs över), för senare
    utvärdering mot faktisk forward-avkastning via --utvardera. claude_rek
    lagras så episodmätningen (§3b) kan splitta på Claudes rek vid inträde.
    """
    claude_texts = claude_texts or {}
    facit = []
    if os.path.exists(FACIT_FILE):
        try:
            with open(FACIT_FILE) as f:
                facit = json.load(f)
        except (json.JSONDecodeError, OSError):
            facit = []
    facit = [r for r in facit if r.get("datum") != datum]   # ersätt dagens rader
    for r in ranking:
        tk = r["ticker"]
        a = analyses.get(tk, {})
        dv = (divergence or {}).get(tk, {})
        facit.append({
            "datum": datum,
            "ticker": tk,
            "poäng": r["poäng"],
            "komponenter": r["delpoäng"],
            "pris": a.get("pris"),
            "viktad_konsensus": consensus.get(tk, {}).get("viktad_konsensus"),
            "divergens_pp": dv.get("divergens_pp"),
            "claude_rek": (claude_texts.get(tk) or {}).get("rekommendation"),
        })
    with open(FACIT_FILE, "w") as f:
        json.dump(facit, f, ensure_ascii=False, indent=2)
    print(f"  Facit uppdaterat: {sum(1 for r in facit if r['datum'] == datum)} rader för {datum} "
          f"(totalt {len(facit)}).")


# ----------------------------------------------------------------------
# §3b Del B: Fyra pappersportföljer — målvikter loggas per körning
# ----------------------------------------------------------------------
CLAUDE_VIKTFAKTOR = {"KÖP": 1.0, "AVVAKTA": 0.5, "SÄLJ": 0.0}


def _normalisera(raw):
    """{tk: råvikt} → {tk: andel} normaliserat till 1.0. Tomt/allt-noll → {}."""
    s = sum(v for v in raw.values() if v > 0)
    if s <= 0:
        return {}
    return {tk: round(v / s, 4) for tk, v in raw.items() if v > 0}


def pappersportfolj_vikter(ranking, claude_texts=None, regim=None, tidigare_ombalanseringar=None):
    """Målvikter för de tre aktiva pappersportföljerna (P3 = SPY, inga vikter).

    Universum = Bästa köp = konsensusaktier med poäng ≥ BASTA_KOP_MIN_POANG
    (exit-aktier ingår aldrig — build_ranking har redan sorterat ut dem).
    Vikter är andelar (summa ≤ 1); resten hålls som 0 %-avkastande kassa.

    - likaviktad (P1): 1/N över universumet — mäter urvalets värde.
    - poangviktad (P2): ∝ max(poäng − 50, 0), normaliserat — mäter poängmodellen.
    - claude (P4): P2:s vikter × Claude-faktor (KÖP 1,0 · AVVAKTA 0,5 · SÄLJ 0);
      den frigjorda vikten går till KASSA (KÖP-aktiernas vikter är oförändrade
      mot P2, så P2 och P4 skiljer sig ENDAST för AVVAKTA/SÄLJ-aktier).
      Saknas rek → behandlas som KÖP (neutral, ingen nedvikt).

    §A regimfilter: i RÖD regim görs INGA nya köp — tickers som inte fanns i
    föregående ombalanserings universum (approximerat via dess likaviktade
    portfölj, som alltid täcker hela universumet) utesluts och räknas i
    "nya_i_kassa" i stället för att viktas.
    """
    claude_texts = claude_texts or {}
    universum = [r["ticker"] for r in ranking
                 if r.get("poäng", 0) >= BASTA_KOP_MIN_POANG]

    nya_i_kassa = []
    if (regim or {}).get("regim") == "RÖD" and tidigare_ombalanseringar:
        senaste = max(tidigare_ombalanseringar, key=lambda d: d["datum"])
        forra_universum = set(senaste.get("portfoljer", {}).get("likaviktad", {}))
        nya_i_kassa = [tk for tk in universum if tk not in forra_universum]
        if nya_i_kassa:
            universum = [tk for tk in universum if tk not in nya_i_kassa]

    if not universum:
        return {"likaviktad": {}, "poangviktad": {}, "claude": {}, "nya_i_kassa": nya_i_kassa}

    likaviktad = {tk: round(1 / len(universum), 4) for tk in universum}

    poäng = {r["ticker"]: r.get("poäng", 0) for r in ranking}
    poangviktad = _normalisera({tk: max(poäng.get(tk, 0) - 50, 0) for tk in universum})

    # P4: applicera Claude-faktorn på P2:s NORMALISERADE vikter, fritt → kassa
    claude = {}
    for tk, v in poangviktad.items():
        rek = (claude_texts.get(tk) or {}).get("rekommendation")
        faktor = CLAUDE_VIKTFAKTOR.get(rek, 1.0)   # okänd rek → neutral (KÖP)
        ny = round(v * faktor, 4)
        if ny > 0:
            claude[tk] = ny
    return {"likaviktad": likaviktad, "poangviktad": poangviktad, "claude": claude,
            "nya_i_kassa": nya_i_kassa}


def logga_pappersportfolj(datum, ranking, claude_texts=None, regim=None):
    """Appenda dagens målvikter till PAPPER_FILE (idempotent — nyckel: datum)."""
    hist = []
    if os.path.exists(PAPPER_FILE):
        try:
            with open(PAPPER_FILE) as f:
                hist = json.load(f)
        except (json.JSONDecodeError, OSError):
            hist = []
    hist = [d for d in hist if d.get("datum") != datum]   # ersätt dagens post (idempotent)
    vikter = pappersportfolj_vikter(ranking, claude_texts, regim, hist)
    post = {"datum": datum, "portfoljer": {k: v for k, v in vikter.items() if k != "nya_i_kassa"}}
    if vikter["nya_i_kassa"]:
        post["nya_i_kassa"] = vikter["nya_i_kassa"]
    hist.append(post)
    hist.sort(key=lambda d: d["datum"])
    with open(PAPPER_FILE, "w") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)
    print(f"  Pappersportföljer: målvikter loggade för {datum} "
          f"({len(vikter['likaviktad'])} aktier, {len(hist)} ombalanseringar totalt).")
    if vikter["nya_i_kassa"]:
        print(f"    regim: RÖD, {len(vikter['nya_i_kassa'])} kandidater hölls i kassa "
              f"({', '.join(vikter['nya_i_kassa'])})")


# ----------------------------------------------------------------------
# §3b: Avkastningsmätning — delade prishjälpare (justerade priser)
# ----------------------------------------------------------------------
def _adj_close(tk, yf, cache):
    """Justerad stängningskurs (auto_adjust=True) som pd.Series, cachad. None vid miss."""
    if tk not in cache:
        try:
            h = yf.Ticker(tk).history(period="2y", auto_adjust=True)
            cache[tk] = h["Close"] if not h.empty else None
        except Exception:
            cache[tk] = None
    return cache[tk]


def _price_asof(close, datum):
    """Pris närmaste handelsdag PÅ eller FÖRE datum (ISO-str). None om serien saknas.

    Datum före seriens start → första tillgängliga pris (best effort).
    """
    import pandas as pd
    if close is None or len(close) == 0:
        return None
    d = pd.Timestamp(datum)
    if close.index.tz is not None:
        d = d.tz_localize(close.index.tz)
    pos = int(close.index.searchsorted(d, side="right")) - 1
    pos = max(0, min(pos, len(close) - 1))
    return float(close.iloc[pos])


# ----------------------------------------------------------------------
# §3b Del A: Episodmätning
# ----------------------------------------------------------------------
UT_ORSAK = {"UT UR KONSENSUS": "konsensus tappad", "EXIT (TRENDBROTT)": "trendbrott"}


def rekonstruera_episoder(history_log):
    """Bygg episoder ur IN-posterna och de två UT-typerna (= Bästa köp-perioder).

    En Bästa köp-episod kan avslutas av antingen konsensusbortfall (UT UR
    KONSENSUS) eller §B:s trendbrott (EXIT (TRENDBROTT)) — vilken som helst
    stänger episoden, ut_orsak skiljer dem åt (§UTBYGGNAD_regim_exit.md).

    Returnerar (episoder, obalanserade_ut). Varje episod: {ticker, in_datum,
    ut_datum|None, ut_orsak|None}. Öppen episod (ingen UT) mäts till idag av
    mät_episoder. UT utan föregående IN (aktier som var konsensus redan i
    första snapshoten) kan inte tidsbestämmas → räknas bara, ingen episod.
    """
    ev = {}
    for e in history_log or []:
        if e["typ"] == "IN I KONSENSUS":
            ev.setdefault(e["ticker"], []).append((e["datum"], "IN", None))
        elif e["typ"] in UT_ORSAK:
            ev.setdefault(e["ticker"], []).append((e["datum"], "UT", UT_ORSAK[e["typ"]]))

    episoder, obalanserade_ut = [], 0
    for tk, lst in ev.items():
        lst.sort(key=lambda x: (x[0], x[1]))   # datum, sen IN < UT samma dag
        öppen = None
        for datum, typ, orsak in lst:
            if typ == "IN":
                if öppen is None:
                    öppen = datum
            else:
                if öppen is not None:
                    episoder.append({"ticker": tk, "in_datum": öppen, "ut_datum": datum,
                                     "ut_orsak": orsak})
                    öppen = None
                else:
                    obalanserade_ut += 1
        if öppen is not None:
            episoder.append({"ticker": tk, "in_datum": öppen, "ut_datum": None, "ut_orsak": None})
    return episoder, obalanserade_ut


def mät_episoder(episoder, facit, yf, sektor_map):
    """Mät varje episods överavkastning mot SPY (+ sektor-ETF om känd).

    Returnerar (rader, obs). Varje rad: ticker, in/ut-datum, dagar, avkastning,
    spy, överavkastning, sektor_etf_överavk, poäng+claude_rek vid inträde,
    öppen (bool). Episoder utan prisdata exkluderas (räknas i obs['exkluderade']).
    """
    from datetime import date
    cache = {}
    idag = date.today().isoformat()

    # facit-uppslag: närmaste rad PÅ eller EFTER inträdesdatumet per ticker
    facit_by_tk = {}
    for r in facit:
        facit_by_tk.setdefault(r["ticker"], []).append(r)
    for lst in facit_by_tk.values():
        lst.sort(key=lambda r: r["datum"])

    def facit_at(tk, datum):
        for r in facit_by_tk.get(tk, []):
            if r["datum"] >= datum:
                return r
        return None

    rader, exkluderade = [], 0
    for ep in episoder:
        tk = ep["ticker"]
        in_d = ep["in_datum"]
        öppen = ep["ut_datum"] is None
        ut_d = ep["ut_datum"] or idag

        close = _adj_close(tk, yf, cache)
        p_in, p_ut = _price_asof(close, in_d), _price_asof(close, ut_d)
        if not p_in or not p_ut:
            exkluderade += 1
            continue
        avk = (p_ut / p_in - 1) * 100

        spy = _adj_close("SPY", yf, cache)
        s_in, s_ut = _price_asof(spy, in_d), _price_asof(spy, ut_d)
        spy_avk = (s_ut / s_in - 1) * 100 if (s_in and s_ut) else None
        överavk = round(avk - spy_avk, 1) if spy_avk is not None else None

        sekt = sektor_map.get(tk) or {}
        etf = _etf_for(sekt.get("sector"), sekt.get("industry")) if sekt else None
        sekt_överavk = None
        if etf and etf != "SPY":
            ec = _adj_close(etf, yf, cache)
            e_in, e_ut = _price_asof(ec, in_d), _price_asof(ec, ut_d)
            if e_in and e_ut:
                sekt_överavk = round(avk - (e_ut / e_in - 1) * 100, 1)

        fr = facit_at(tk, in_d) or {}
        rader.append({
            "ticker": tk, "in_datum": in_d, "ut_datum": ep["ut_datum"], "öppen": öppen,
            "ut_orsak": ep.get("ut_orsak"),
            "dagar": (date.fromisoformat(ut_d) - date.fromisoformat(in_d)).days,
            "avkastning": round(avk, 1),
            "spy": round(spy_avk, 1) if spy_avk is not None else None,
            "överavkastning": överavk,
            "sektor_etf_överavk": sekt_överavk, "sektor_etf": etf,
            "poäng_in": fr.get("poäng"), "claude_rek_in": fr.get("claude_rek"),
        })
    return rader, {"exkluderade": exkluderade}


# ----------------------------------------------------------------------
# §3b Del B: Pappersportföljernas avkastning (kedjade perioder)
# ----------------------------------------------------------------------
def utvardera_pappersportfoljer(yf):
    """Kedja ihop pappersportföljernas avkastning mellan ombalanseringarna.

    Returnerar dict med per-portfölj-nyckeltal (totalavkastning, snittomsättning,
    max drawdown), SPY-benchmark, parvisa differenser och n — eller None om
    färre än 2 ombalanseringar finns.
    """
    if not os.path.exists(PAPPER_FILE):
        return None
    try:
        with open(PAPPER_FILE) as f:
            hist = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    hist = sorted(hist, key=lambda d: d["datum"])
    if len(hist) < 2:
        return None

    from datetime import date
    cache = {}
    datum = [d["datum"] for d in hist]
    perioder = list(zip(datum, datum[1:] + [date.today().isoformat()]))   # sista → idag
    portnamn = ["likaviktad", "poangviktad", "claude"]

    def periodavk(vikter, d0, d1):
        r = 0.0
        for tk, w in vikter.items():
            c = _adj_close(tk, yf, cache)
            p0, p1 = _price_asof(c, d0), _price_asof(c, d1)
            if p0 and p1:
                r += w * (p1 / p0 - 1)   # kassa (1−Σw) bidrar 0
        return r

    def maxdrawdown(kurva):
        topp, mdd = kurva[0], 0.0
        for v in kurva:
            topp = max(topp, v)
            mdd = min(mdd, v / topp - 1)
        return round(mdd * 100, 1)

    resultat = {}
    for p in portnamn:
        kurva = [1.0]
        for (d0, d1) in perioder:
            kurva.append(kurva[-1] * (1 + periodavk(hist[datum.index(d0)]["portfoljer"][p], d0, d1)))
        oms = [sum(abs(hist[i]["portfoljer"][p].get(t, 0) - hist[i - 1]["portfoljer"][p].get(t, 0))
                   for t in set(hist[i]["portfoljer"][p]) | set(hist[i - 1]["portfoljer"][p])) / 2
               for i in range(1, len(hist))]
        resultat[p] = {
            "totalavkastning": round((kurva[-1] - 1) * 100, 1),
            "snittomsattning": round(sum(oms) / len(oms) * 100, 1) if oms else 0.0,
            "max_drawdown": maxdrawdown(kurva),
        }

    # P3: SPY buy-and-hold över hela perioden (datum[0] → idag)
    spy = _adj_close("SPY", yf, cache)
    s0, s1 = _price_asof(spy, datum[0]), _price_asof(spy, date.today().isoformat())
    spy_tot = round((s1 / s0 - 1) * 100, 1) if (s0 and s1) else None
    resultat["benchmark_spy"] = {"totalavkastning": spy_tot}

    diffar = []
    p1, p2, p4 = (resultat["likaviktad"]["totalavkastning"],
                  resultat["poangviktad"]["totalavkastning"],
                  resultat["claude"]["totalavkastning"])
    if spy_tot is not None:
        diffar.append(("Urvalets värde (P1 − SPY)", round(p1 - spy_tot, 1)))
    diffar.append(("Poängmodellens värde (P2 − P1)", round(p2 - p1, 1)))
    diffar.append(("Claude-rekommendationens värde (P4 − P2)", round(p4 - p2, 1)))

    return {"resultat": resultat, "diffar": diffar, "n": len(perioder),
            "start": datum[0], "slut": date.today().isoformat()}


def run_utvardering():
    """Utvärdera poängmodellen mot faktisk forward-avkastning (21/63/126 dgr).

    Läser FACIT_FILE, hämtar prishistorik via yfinance och rapporterar
    avkastning per poängkvartil, komponentkorrelationer och divergensutfall
    till terminal + Excel-fliken 'Utvärdering' (utvardering.xlsx). Kör INTE
    analysen — glesa datapunkter är OK, rapporten anger n.
    """
    import yfinance as yf
    import pandas as pd

    gist_pull()
    if not os.path.exists(FACIT_FILE):
        raise RuntimeError(f"{FACIT_FILE} saknas — kör analysen minst en gång först.")
    with open(FACIT_FILE) as f:
        facit = json.load(f)
    if not facit:
        raise RuntimeError("Facit är tomt — kör analysen några gånger på olika datum först.")

    HORISONTER = [("21d", 21), ("63d", 63), ("126d", 126)]
    PRIMÄR = "63d"
    tickers = sorted({r["ticker"] for r in facit})
    BLIND_FLACK = (
        "Utvärdering av intern rangordning (mäter ej täckning) — facit "
        "innehåller bara aktier screenern redan lyfte fram. En bra "
        "kvartilkorrelation visar att modellen rangordnar KANDIDATERNA väl, "
        "inte att den hittar de bästa aktierna på hela marknaden."
    )
    print(BLIND_FLACK)
    print(f"\nUtvärderar {len(facit)} facitrader, {len(tickers)} unika aktier...")

    close_cache = {}
    for tk in tickers:
        try:
            h = yf.Ticker(tk).history(period="2y")
            if not h.empty:
                close_cache[tk] = h["Close"]
        except Exception:
            pass

    # Forward-avkastning per rad (närmaste handelsdag ≥ raddatumet)
    for r in facit:
        r["fwd"] = {}
        close = close_cache.get(r["ticker"])
        if close is None:
            continue
        d = pd.Timestamp(r["datum"])
        if close.index.tz is not None:
            d = d.tz_localize(close.index.tz)
        pos = int(close.index.searchsorted(d))
        if pos >= len(close):
            continue
        p0 = float(close.iloc[pos])
        for namn, n in HORISONTER:
            if pos + n < len(close) and p0:
                r["fwd"][namn] = round((float(close.iloc[pos + n]) / p0 - 1) * 100, 1)

    # Datamängd med primärhorisontens forward-avkastning
    rader = [r for r in facit if r["fwd"].get(PRIMÄR) is not None]
    n = len(rader)
    print(f"\nDatapunkter med {PRIMÄR} forward-avkastning: {n}")
    if n == 0:
        print("Ännu inga färdiga forward-fönster — poängkvartilerna hoppas över, "
              "men episod- och pappersmätningen körs ändå.")
    elif n < 10:
        print(f"VARNING: för tidigt för slutsatser (n={n}). Behöver ~10+ datapunkter.")

    rapport = []   # (rubrik, [(etikett, värde)]) — forward-fönstrets del
    if n > 0:
        df = pd.DataFrame([{
            "poäng": r["poäng"], "divergens_pp": r.get("divergens_pp"),
            "fwd": r["fwd"][PRIMÄR],
            **{f"k_{k}": v for k, v in (r.get("komponenter") or {}).items()},
        } for r in rader])

        # 1. Avkastning per poängkvartil
        kvartil_rader = []
        try:
            df["kvartil"] = pd.qcut(df["poäng"], 4, labels=["Q1 (lägst)", "Q2", "Q3", "Q4 (högst)"],
                                    duplicates="drop")
            for kv, grp in df.groupby("kvartil", observed=True):
                kvartil_rader.append((str(kv), f"snitt {grp['fwd'].mean():+.1f} %  (n={len(grp)})"))
        except (ValueError, IndexError):
            kvartil_rader.append(("—", "för få unika poäng för kvartiler"))
        rapport.append((f"Forward-avkastning ({PRIMÄR}) per poängkvartil", kvartil_rader))

        # 2. Komponentkorrelation mot forward-avkastning
        komp_rader = []
        for kol in [c for c in df.columns if c.startswith("k_")]:
            if df[kol].nunique() > 1:
                korr = df[kol].corr(df["fwd"])
                komp_rader.append((kol[2:], f"korr {korr:+.2f}"))
        komp_rader.append(("Totalpoäng", f"korr {df['poäng'].corr(df['fwd']):+.2f}"))
        rapport.append(("Korrelation komponentpoäng ↔ forward-avkastning", komp_rader))

        # 3. Divergensutfall (hög vs låg divergens)
        div_rader = []
        dd = df.dropna(subset=["divergens_pp"])
        if len(dd) >= 4:
            median = dd["divergens_pp"].median()
            hög = dd[dd["divergens_pp"] >= median]["fwd"].mean()
            låg = dd[dd["divergens_pp"] < median]["fwd"].mean()
            div_rader = [(f"Hög divergens (≥ {median:.0f} pp)", f"snitt {hög:+.1f} %"),
                         (f"Låg divergens (< {median:.0f} pp)", f"snitt {låg:+.1f} %")]
        else:
            div_rader = [("—", "för få divergensdatapunkter")]
        rapport.append(("Divergensutfall", div_rader))

        # Terminalrapport (forward-fönstret)
        for rubrik, poster in rapport:
            print(f"\n{rubrik}:")
            for etikett, värde in poster:
                print(f"  {etikett:<22} {värde}")

    # ---- §3b Del A: Episodmätning ----
    from statistics import mean, median
    sektor_map = {}
    if os.path.exists(RESULTS_FILE):
        try:
            with open(RESULTS_FILE) as f:
                for tk, a in (json.load(f).get("analyses") or {}).items():
                    if isinstance(a, dict):
                        sektor_map[tk] = {"sector": a.get("sector"), "industry": a.get("industry")}
        except (json.JSONDecodeError, OSError):
            pass
    history_log = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE) as f:
                history_log = json.load(f).get("logg") or []
        except (json.JSONDecodeError, OSError):
            pass

    episoder, obalanserade = rekonstruera_episoder(history_log)
    ep_rader, ep_obs = mät_episoder(episoder, facit, yf, sektor_map)
    stängda = [r for r in ep_rader if not r["öppen"] and r["överavkastning"] is not None]
    öppna = [r for r in ep_rader if r["öppen"] and r["överavkastning"] is not None]

    print("\n\n=== §3b Del A: Episodmätning (Bästa köp-perioder mot SPY) ===")
    print(f"{len(ep_rader)} episoder mätta ({len(stängda)} stängda, {len(öppna)} öppna); "
          f"{ep_obs['exkluderade']} utan prisdata, {obalanserade} UT utan känt inträde.")

    def episod_sammanfattning(rader, etikett):
        if not rader:
            print(f"  {etikett}: inga episoder.")
            return
        öa = [r["överavkastning"] for r in rader]
        träff = sum(1 for x in öa if x > 0) / len(öa) * 100
        varn = "  ⚠ för tidigt för slutsatser" if len(rader) < 10 else ""
        print(f"  {etikett} (n={len(rader)}):{varn}")
        print(f"    snittöveravkastning {mean(öa):+.1f} pp · median {median(öa):+.1f} pp · "
              f"träffprocent {träff:.0f} %")

    print("\n  STÄNGDA episoder (facit — realiserat utfall):")
    episod_sammanfattning(stängda, "Alla stängda")
    episod_sammanfattning([r for r in stängda if r["claude_rek_in"] == "KÖP"], "  varav Claude KÖP")
    episod_sammanfattning([r for r in stängda if r["claude_rek_in"] == "AVVAKTA"], "  varav Claude AVVAKTA")
    episod_sammanfattning([r for r in stängda if r["ut_orsak"] == "trendbrott"], "  varav ut: trendbrott (§B)")
    episod_sammanfattning([r for r in stängda if r["ut_orsak"] == "konsensus tappad"],
                          "  varav ut: konsensus tappad")
    print("\n  ÖPPNA episoder (SEPARAT — kvarliggare har överlevnadsfel, blanda ej):")
    episod_sammanfattning(öppna, "Alla öppna")

    # ---- §3b Del B: Pappersportföljer ----
    pp = utvardera_pappersportfoljer(yf)
    print("\n\n=== §3b Del B: Fyra pappersportföljer ===")
    if pp is None:
        print("  För få ombalanseringar än (behöver ≥ 2 körningar på olika datum).")
    else:
        NAMN = {"likaviktad": "P1 Likaviktad", "poangviktad": "P2 Poängviktad",
                "claude": "P4 Poäng+Claude"}
        varn = "  ⚠ för tidigt (n={})".format(pp["n"]) if pp["n"] < 8 else ""
        print(f"  {pp['start']} → {pp['slut']}, {pp['n']} ombalanseringsperioder.{varn}")
        for nyckel, namn in NAMN.items():
            r = pp["resultat"][nyckel]
            print(f"    {namn:<18} totalavk {r['totalavkastning']:+6.1f} % · "
                  f"omsättn {r['snittomsattning']:.0f} %/ombal · maxDD {r['max_drawdown']:.1f} %")
        spy = pp["resultat"]["benchmark_spy"]["totalavkastning"]
        print(f"    {'P3 SPY (b&h)':<18} totalavk {spy:+6.1f} %" if spy is not None
              else f"    P3 SPY: prisdata saknas")
        print("  Parvisa differenser (vad varje lager tillför):")
        for etikett, v in pp["diffar"]:
            print(f"    {etikett:<40} {v:+.1f} pp")

    # ---- Claude-förbrukning (terminalrapport, ingen Excel-flik) ----
    # Uppdelad i normal drift (genuin trigger) och felsökning/force (bara
    # --force-claude, ingen egen orsak) så testkörningar inte förorenar
    # driftssiffrorna — se _är_force_post.
    fr = claude_forbrukning_rapport()
    print("\n\n=== Claude-förbrukning ===")
    if fr is None:
        print("  Ingen förbrukningslogg än (claude_forbrukning.json saknas/tom).")
    else:
        drift, force = fr["drift"], fr["force"]
        if drift["total_antal"]:
            print(f"  Normal drift: {drift['total_antal']} anrop, {drift['total_tokens']} tokens "
                  f"(snitt {drift['snitt_per_anrop']} tokens/anrop).")
        else:
            print("  Normal drift: inga anrop loggade än.")
        if force["total_antal"]:
            print(f"    + {force['total_antal']} force-anrop under felsökning "
                  f"({force['total_tokens']} tokens, snitt {force['snitt_per_anrop']} tokens/anrop)")

        if drift["total_antal"]:
            print("  Per vecka (senaste 8, normal drift):")
            for vecka in sorted(drift["per_vecka"], reverse=True)[:8]:
                v = drift["per_vecka"][vecka]
                print(f"    {vecka}   {v['antal_anrop']:>3} anrop   {v['tokens']:>8} tokens")
            print("  Per orsak (normal drift — vad triggar mest, sorterat på tokens):")
            for orsak, v in sorted(drift["per_orsak"].items(), key=lambda x: -x[1]["tokens"]):
                snitt = round(v["tokens"] / v["antal"]) if v["antal"] else 0
                print(f"    {orsak:<40} {v['antal']:>3} anrop   {v['tokens']:>8} tokens   (snitt {snitt}/anrop)")
            print("  Per modell (normal drift):")
            for modell, v in sorted(drift["per_modell"].items(), key=lambda x: -x[1]["tokens"]):
                snitt = round(v["tokens"] / v["antal"]) if v["antal"] else 0
                print(f"    {modell:<10} {v['antal']:>3} anrop   {v['tokens']:>8} tokens   (snitt {snitt}/anrop)")

    # Excel-flik
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Utvärdering"
    ws.append([f"Utvärdering av intern rangordning (mäter ej täckning) — "
               f"{n} datapunkter ({PRIMÄR} forward)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([BLIND_FLACK])
    if n < 10:
        ws.append([f"VARNING: för tidigt för slutsatser (n={n})"])
    for rubrik, poster in rapport:
        ws.append([])
        ws.append([rubrik])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for etikett, värde in poster:
            ws.append([etikett, värde])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30

    # Excel-flik: Episoder (Del A)
    wsE = wb.create_sheet("Episoder")
    wsE.append(["Episodmätning — Bästa köp-perioder mot SPY (justerade priser). "
                "Överavkastning = episod − SPY samma intervall."])
    wsE["A1"].font = Font(bold=True, size=12)
    if len(stängda) < 10:
        wsE.append([f"VARNING: för tidigt för slutsatser (n={len(stängda)} stängda)"])
    wsE.append([])
    wsE.append(["Ticker", "In-datum", "Ut-datum", "Ut-orsak", "Dagar", "Avkastning (%)",
                "SPY samma period (%)", "Överavkastning (pp)",
                "Sektor-ETF-överavk (pp)", "Poäng vid inträde", "Claude-rek vid inträde"])
    for c in wsE[wsE.max_row]:
        c.font = Font(bold=True)
    for r in sorted(ep_rader, key=lambda x: (x["öppen"], -(x["överavkastning"] or -999))):
        wsE.append([r["ticker"], r["in_datum"], r["ut_datum"] or "öppen", r.get("ut_orsak"), r["dagar"],
                    r["avkastning"], r["spy"], r["överavkastning"],
                    r["sektor_etf_överavk"], r["poäng_in"], r["claude_rek_in"]])
    for kol, br in zip("ABCDEFGHIJK", (10, 12, 12, 16, 7, 14, 18, 18, 18, 16, 20)):
        wsE.column_dimensions[kol].width = br

    # Excel-flik: Pappersportföljer (Del B)
    wsP = wb.create_sheet("Pappersportföljer")
    wsP.append(["Fyra pappersportföljer — kedjad avkastning över ombalanseringar "
                "(justerade priser, ingen transaktionskostnad)."])
    wsP["A1"].font = Font(bold=True, size=12)
    if pp is None:
        wsP.append(["För få ombalanseringar än (behöver ≥ 2 körningar på olika datum)."])
    else:
        if pp["n"] < 8:
            wsP.append([f"VARNING: för tidigt (n={pp['n']} perioder)"])
        wsP.append([f"Period: {pp['start']} → {pp['slut']}  ·  {pp['n']} ombalanseringar"])
        wsP.append([])
        wsP.append(["Portfölj", "Totalavkastning (%)", "Snittomsättning (%/ombal)", "Max drawdown (%)"])
        for c in wsP[wsP.max_row]:
            c.font = Font(bold=True)
        NAMN = {"likaviktad": "P1 Likaviktad", "poangviktad": "P2 Poängviktad",
                "claude": "P4 Poäng + Claude-filter"}
        for nyckel, namn in NAMN.items():
            r = pp["resultat"][nyckel]
            wsP.append([namn, r["totalavkastning"], r["snittomsattning"], r["max_drawdown"]])
        spy = pp["resultat"]["benchmark_spy"]["totalavkastning"]
        wsP.append(["P3 SPY (buy & hold)", spy, "—", "—"])
        wsP.append([])
        wsP.append(["Parvis differens (vad lagret tillför)", "Skillnad (pp)"])
        wsP.cell(row=wsP.max_row, column=1).font = Font(bold=True)
        for etikett, v in pp["diffar"]:
            wsP.append([etikett, v])
        for kol, br in zip("ABCD", (38, 22, 26, 18)):
            wsP.column_dimensions[kol].width = br

    wb.save("utvardering.xlsx")
    print("\nRapport sparad som: utvardering.xlsx (flikar: Utvärdering, Episoder, Pappersportföljer)")


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="eToro portföljanalys — konsensus, teknisk analys och divergens.",
        epilog="Utan flaggor körs standardanalysen av signalgruppens 5 profiler "
               "(divergensen räknas från senast cachade bakgrundsportföljer).")
    parser.add_argument("--screener", action="store_true",
                        help=f"screena fram bakgrundsgruppen (topp {BG_SIZE}) och spara till "
                             f"{BG_MEMBERS_FILE} — kör inte analysen")
    parser.add_argument("--divergens", action="store_true",
                        help="hämta om bakgrundsgruppens portföljer (kräver att --screener "
                             "körts någon gång) och kör sedan hela analysen")
    parser.add_argument("--force-claude", action="store_true",
                        help="kör Claude-analysen även om den redan körts idag (drar credits)")
    parser.add_argument("--utvardera", action="store_true",
                        help="utvärdera poängmodellen mot faktisk forward-avkastning "
                             f"({FACIT_FILE}) — kör inte analysen")
    args = parser.parse_args()

    print("=== eToro portföljanalys ===")
    try:
        if args.utvardera:
            run_utvardering()
            return
        if args.screener:
            if not API_KEY or not USER_KEY:
                raise RuntimeError("ETORO_API_KEY och ETORO_USER_KEY saknas — lägg dem i .env-filen.")
            gist_pull()
            if run_screener() is None:
                raise RuntimeError("Screenern misslyckades — se utskriften ovan.")
            gist_push()
            return
        run_analysis(force_claude=args.force_claude, refresh_background=args.divergens)
    except RuntimeError as e:
        sys.exit(f"\nFEL: {e}")


if __name__ == "__main__":
    main()
